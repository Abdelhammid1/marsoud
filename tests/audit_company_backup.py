#!/usr/bin/env python3
"""MARSOUD — audit the full-company Excel backup feature.

Covers:
  - build_company_workbook returns a non-empty BytesIO
  - workbook is readable by openpyxl + has the expected sheet set
  - chart-of-accounts sheet has at least one data row
  - filename helper produces a safe + timestamped name
  - HTTP: owner reaches /settings/backup/ (200)
  - HTTP: POST /settings/backup/excel returns the xlsx mimetype
  - HTTP: non-owner is redirected (302) — gate works
  - HTTP: unauthenticated is redirected to /login
"""
import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app import create_app, db


CHECKS = []


def check(label):
    def deco(fn):
        CHECKS.append((label, fn))
        return fn
    return deco


def _fixture():
    from app.models import Company, User
    return {
        "company": Company.query.first(),
        "owner": User.query.filter_by(email="demo@manasety.ai").first(),
    }


# ─── Service layer ──────────────────────────────────────────────────────
@check("1. build_company_workbook returns non-empty BytesIO")
def _():
    from app.services.company_backup import build_company_workbook
    f = _fixture()
    buf = build_company_workbook(f["company"].id)
    n = len(buf.getvalue())
    assert isinstance(buf, io.BytesIO), f"got {type(buf)}"
    assert n > 5000, f"workbook suspiciously small: {n} bytes"
    return f"buffer: {n} bytes"


@check("2. workbook contains the expected core sheets")
def _():
    from app.services.company_backup import build_company_workbook
    from openpyxl import load_workbook
    f = _fixture()
    buf = build_company_workbook(f["company"].id)
    wb = load_workbook(buf, read_only=True, data_only=True)
    must_have = {"الشركة", "شجرة الحسابات", "القيود اليومية",
                  "العملاء", "الموردين", "المنتجات",
                  "فواتير المبيعات", "الموظفين"}
    missing = must_have - set(wb.sheetnames)
    assert not missing, f"missing sheets: {missing}"
    return f"{len(wb.sheetnames)} sheets, all core sheets present"


@check("3. شجرة الحسابات sheet has the right header + at least 1 data row")
def _():
    from app.services.company_backup import build_company_workbook
    from openpyxl import load_workbook
    f = _fixture()
    buf = build_company_workbook(f["company"].id)
    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb["شجرة الحسابات"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert headers[:3] == ["كود", "الاسم", "الاسم بالعربية"], \
        f"unexpected headers: {headers}"
    assert ws.max_row >= 2, f"only header row present (max_row={ws.max_row})"
    first_data = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    assert first_data[0], f"first data row has no code: {first_data}"
    return f"{ws.max_row - 1} accounts exported"


@check("4. filename helper is safe + includes a timestamp")
def _():
    from app.services.company_backup import workbook_filename
    f = _fixture()
    fn = workbook_filename(f["company"])
    assert fn.startswith("marsoud-backup-"), fn
    assert fn.endswith(".xlsx"), fn
    # No path traversal characters
    assert "/" not in fn and "\\" not in fn and ".." not in fn, fn
    return fn


@check("5. ValueError on unknown company id")
def _():
    from app.services.company_backup import build_company_workbook
    try:
        build_company_workbook(999999)
    except ValueError as e:
        assert "Company" in str(e) or "999999" in str(e)
        return f"raised: {e}"
    raise AssertionError("did not raise for missing company")


# ─── HTTP layer ─────────────────────────────────────────────────────────
@check("6. /settings/backup/ for owner returns 200 + page renders")
def _():
    app = create_app()
    with app.test_client() as c:
        c.post("/login", data={"email": "demo@manasety.ai",
                                 "password": "demo1234"})
        r = c.get("/settings/backup/")
        assert r.status_code == 200, f"status={r.status_code}"
        html = r.get_data(as_text=True)
        assert "نسخة احتياطية" in html, "page title text missing"
    return "owner reaches page"


@check("7. POST /settings/backup/excel returns the xlsx mimetype")
def _():
    app = create_app()
    with app.test_client() as c:
        c.post("/login", data={"email": "demo@manasety.ai",
                                 "password": "demo1234"})
        r = c.post("/settings/backup/excel")
        assert r.status_code == 200, f"status={r.status_code}"
        assert r.mimetype == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ), f"mimetype={r.mimetype}"
        # Content-Disposition includes UTF-8 filename* form for Arabic
        cd = r.headers.get("Content-Disposition", "")
        assert "attachment" in cd, f"not attachment: {cd}"
        assert "xlsx" in cd, f"no xlsx in CD: {cd}"
        assert len(r.data) > 5000, f"body too small: {len(r.data)}"
    return f"xlsx download OK ({len(r.data)} bytes)"


@check("8. Anonymous gets redirected to /login")
def _():
    app = create_app()
    with app.test_client() as c:
        r = c.get("/settings/backup/", follow_redirects=False)
        assert r.status_code in (302, 303), f"status={r.status_code}"
        loc = r.headers.get("Location", "")
        assert "login" in loc, f"redirect to {loc}, expected /login"
    return f"anon → {loc}"


@check("9. Non-owner cannot reach the page or download")
def _():
    """Build a viewer-role user, log them in, both URLs must redirect away."""
    from werkzeug.security import generate_password_hash
    from app.models import User, Company
    from app.models.user import user_companies
    EMAIL = "backup_audit_viewer@example.com"
    PW = "viewer1234"
    f = _fixture()
    u = User.query.filter_by(email=EMAIL).first()
    if not u:
        u = User(email=EMAIL, full_name="backup audit viewer",
                 password_hash=generate_password_hash(PW, method="pbkdf2:sha256"),
                 is_active=True)
        db.session.add(u); db.session.flush()
    else:
        u.password_hash = generate_password_hash(PW, method="pbkdf2:sha256")
    row = db.session.execute(user_companies.select().where(
        (user_companies.c.user_id == u.id) &
        (user_companies.c.company_id == f["company"].id)
    )).first()
    if not row:
        db.session.execute(user_companies.insert().values(
            user_id=u.id, company_id=f["company"].id, role="viewer",
        ))
    db.session.commit()
    app = create_app()
    try:
        with app.test_client() as c:
            c.post("/login", data={"email": EMAIL, "password": PW})
            r = c.get("/settings/backup/", follow_redirects=False)
            assert r.status_code in (302, 303), f"page status={r.status_code}"
            r2 = c.post("/settings/backup/excel", follow_redirects=False)
            assert r2.status_code in (302, 303), \
                f"download status={r2.status_code}"
    finally:
        db.session.execute(user_companies.delete().where(
            (user_companies.c.user_id == u.id) &
            (user_companies.c.company_id == f["company"].id)
        ))
        db.session.delete(db.session.get(User, u.id))
        db.session.commit()
    return f"viewer blocked from page + download"


# ─── Run ────────────────────────────────────────────────────────────────
def main():
    app = create_app()
    passed = failed = 0
    failures = []
    with app.app_context():
        for label, fn in CHECKS:
            try:
                result = fn()
                print(f"PASS  {label}  ⇒ {result}")
                passed += 1
            except Exception as e:
                print(f"FAIL  {label}  ⇒ {type(e).__name__}: {e}")
                failures.append((label, repr(e)))
                failed += 1
    print()
    print(f"────  {passed} passed, {failed} failed  ────")
    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
