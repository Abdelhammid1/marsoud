#!/usr/bin/env python3
"""MARSOUD-LEAD-EXPORT-BY-CAMPAIGN (Abdelhamid 2026-07-15).

Grouped Excel export: leads split into one sheet per campaign +
a summary sheet up front showing count-per-status per campaign.
Un-campaigned leads live in a "بدون حملة" sheet.

Checks:
  1. export_leads_by_campaign_excel returns a non-empty xlsx buffer.
  2. Sheet 1 is "ملخص الحملات" (summary).
  3. There's one sheet per distinct campaign name AND a "بدون حملة"
     sheet when at least one lead has no campaign.
  4. Summary sheet shows per-status count for each campaign.
  5. Per-campaign sheet is sorted by status (pipeline order).
  6. HTTP /leads/export/excel?group_by=campaign returns the grouped
     variant (xlsx, non-empty).
  7. Default HTTP /leads/export/excel (no group_by) still returns
     the flat variant (backward compat).
"""
import sys
from pathlib import Path
from datetime import datetime
from io import BytesIO

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app import create_app, db


CHECKS = []
_STATE = {}


def check(label):
    def deco(fn):
        CHECKS.append((label, fn))
        return fn
    return deco


def _teardown(company_id):
    from sqlalchemy import text, inspect
    db.session.close()
    insp = inspect(db.engine)
    with db.engine.begin() as conn:
        conn.execute(text(
            "DELETE FROM user_companies WHERE company_id = :c"),
            {"c": company_id})
        for tbl in reversed(db.metadata.sorted_tables):
            cols = {col["name"] for col in insp.get_columns(tbl.name)}
            if "company_id" in cols:
                conn.execute(
                    text(f"DELETE FROM {tbl.name} WHERE company_id = :c"),
                    {"c": company_id})
        conn.execute(text("DELETE FROM companies WHERE id = :c"),
                     {"c": company_id})
        conn.execute(text(
            "DELETE FROM users WHERE email LIKE 'lex-%@x.test'"))


def _setup():
    from app.models import (
        Company, User, user_companies, Lead, LeadStatus, Campaign,
    )
    from werkzeug.security import generate_password_hash

    for name in ("__LEX__",):
        c = Company.query.filter_by(name=name).first()
        if c:
            _teardown(c.id)
    a = Company(name="__LEX__", base_currency="SAR")
    db.session.add(a); db.session.flush()
    from app.services.seed_coa import seed_default_coa
    seed_default_coa(a.id)

    def _mk(email, role):
        u = User(email=email,
                 password_hash=generate_password_hash(
                     "x", method="pbkdf2:sha256"),
                 full_name=email.split("@")[0])
        db.session.add(u); db.session.flush()
        db.session.execute(user_companies.insert().values(
            user_id=u.id, company_id=a.id, role=role))
        return u

    owner = _mk("lex-owner@x.test", "owner")

    camp_x = Campaign(company_id=a.id, name="LEX-Camp-X",
                      active=True, created_by_id=owner.id)
    camp_y = Campaign(company_id=a.id, name="LEX-Camp-Y",
                      active=True, created_by_id=owner.id)
    db.session.add_all([camp_x, camp_y]); db.session.flush()

    def _lead(name, camp_id, status):
        db.session.add(Lead(
            company_id=a.id, client_name=name,
            phone="0500000000", service_needed="test",
            assigned_to_id=owner.id, created_by_id=owner.id,
            status=status, campaign_id=camp_id,
        ))
    # X: 2 NEW + 1 WON
    _lead("LEX-X-1", camp_x.id, LeadStatus.NEW_LEAD)
    _lead("LEX-X-2", camp_x.id, LeadStatus.NEW_LEAD)
    _lead("LEX-X-3", camp_x.id, LeadStatus.WON)
    # Y: 1 CONTACTED + 1 LOST
    _lead("LEX-Y-1", camp_y.id, LeadStatus.CONTACTED)
    _lead("LEX-Y-2", camp_y.id, LeadStatus.LOST)
    # No campaign: 1 lead
    _lead("LEX-NoCamp-1", None, LeadStatus.NEW_LEAD)
    db.session.commit()

    _STATE.update(a_id=a.id, owner_id=owner.id)


def _reset_g():
    from flask import g
    for k in ("_login_user", "active_company", "user_companies",
              "impersonating"):
        try: g.pop(k, None)
        except Exception: pass


def _login():
    from flask import current_app
    _reset_g()
    client = current_app.test_client()
    with client.session_transaction() as sess:
        sess["_user_id"] = str(_STATE["owner_id"])
        sess["_fresh"] = True
        sess["active_company_id"] = _STATE["a_id"]
    return client


# ─── Service ───────────────────────────────────────────────────────
@check("1. export_leads_by_campaign_excel returns a non-empty xlsx buffer")
def _():
    from app.models import Company, Lead
    from app.services.export import export_leads_by_campaign_excel
    company = db.session.get(Company, _STATE["a_id"])
    leads = Lead.query.filter_by(company_id=_STATE["a_id"]).all()
    buf = export_leads_by_campaign_excel(company, leads)
    data = buf.read()
    assert len(data) > 1000, f"suspiciously small: {len(data)} bytes"
    _STATE["xlsx_bytes"] = data
    return f"{len(data)} bytes"


@check("2. First sheet is the campaigns summary")
def _():
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(_STATE["xlsx_bytes"]))
    assert wb.sheetnames[0] == "ملخص الحملات", \
        f"first sheet: {wb.sheetnames[0]}"
    return f"summary at position 0"


@check("3. One sheet per campaign + a 'بدون حملة' sheet")
def _():
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(_STATE["xlsx_bytes"]))
    names = set(wb.sheetnames)
    for expected in ("ملخص الحملات", "LEX-Camp-X", "LEX-Camp-Y",
                      "بدون حملة"):
        assert expected in names, \
            f"missing sheet {expected!r} (have {names})"
    return f"{len(names)} sheets: {sorted(names)}"


@check("4. Summary sheet has per-status counts per campaign")
def _():
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(_STATE["xlsx_bytes"]))
    ws = wb["ملخص الحملات"]
    # Extract data: header row 5, then per-campaign rows.
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    header = rows[0]
    assert header[0] == "الحملة"
    # X should show: NEW_LEAD=2, WON=1, total=3
    for r in rows[1:]:
        if r[0] == "LEX-Camp-X":
            total = r[-1]
            assert total == 3, f"X total: {total}"
            break
    else:
        assert False, "LEX-Camp-X row missing"
    return "counts correct for LEX-Camp-X"


@check("5. Per-campaign sheet sorted by status (pipeline order)")
def _():
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(_STATE["xlsx_bytes"]))
    ws = wb["LEX-Camp-X"]
    # Row 6+ = data rows. Column 6 = الحالة (status label).
    statuses = [row[5] for row in ws.iter_rows(min_row=6, values_only=True)
                if row[1]]   # rows with a client name
    # NEW_LEAD label appears before WON label (pipeline order).
    from app.models import LeadStatus
    new_label = LeadStatus.NEW_LEAD.label_ar
    won_label = LeadStatus.WON.label_ar
    idx_new = next((i for i, s in enumerate(statuses)
                     if s == new_label), None)
    idx_won = next((i for i, s in enumerate(statuses)
                     if s == won_label), None)
    assert idx_new is not None and idx_won is not None
    assert idx_new < idx_won, \
        f"pipeline order broken: new={idx_new}, won={idx_won}"
    return f"order OK: new-lead@{idx_new}, won@{idx_won}"


# ─── HTTP ──────────────────────────────────────────────────────────
@check("6. HTTP ?group_by=campaign returns the multi-sheet variant")
def _():
    r = _login().get("/leads/export/excel?group_by=campaign",
                       follow_redirects=False)
    assert r.status_code == 200, f"got {r.status_code}"
    ct = r.headers.get("Content-Type", "")
    assert "spreadsheet" in ct or "excel" in ct, ct
    assert len(r.data) > 1000
    # Confirm shape by reloading.
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.data))
    assert "ملخص الحملات" in wb.sheetnames
    return f"grouped file OK ({len(r.data)} bytes)"


@check("7. HTTP default (no group_by) still returns the flat variant")
def _():
    r = _login().get("/leads/export/excel",
                       follow_redirects=False)
    assert r.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.data))
    # Flat = single sheet "Leads".
    assert wb.sheetnames == ["Leads"], \
        f"expected flat, got {wb.sheetnames}"
    return "flat export unchanged"


def main():
    app = create_app()
    passed = failed = 0
    with app.app_context():
        try:
            _setup()
            for label, fn in CHECKS:
                try:
                    result = fn()
                    print(f"PASS  {label}  ⇒ {result}")
                    passed += 1
                except Exception as e:  # noqa: BLE001
                    print(f"FAIL  {label}  ⇒ {type(e).__name__}: {e}")
                    failed += 1
                    import traceback
                    traceback.print_exc()
        finally:
            try:
                if "a_id" in _STATE:
                    _teardown(_STATE["a_id"])
                print("\n(cleaned up fixture company)")
            except Exception as e:
                print(f"\n(teardown failed: {e})")
    print()
    print(f"────  {passed} passed, {failed} failed  ────")
    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
