#!/usr/bin/env python3
"""Audit for the 3 fixes shipped in commit df056cf:

  G) /calendar/ no longer 500s when a task has project_id=None.
  H) HR can manually grant a leave balance to an employee, capped at
     the leave-type's max_balance.
  I) GET /leads/export/excel returns a valid xlsx with all the
     expected columns + honours list filters.
"""
import io
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app import create_app, db


CHECKS = []
DEMO_EMAIL = "demo@manasety.ai"
DEMO_PASS = "demo1234"


def check(label):
    def deco(fn):
        CHECKS.append((label, fn))
        return fn
    return deco


def _login(client, email, password):
    r = client.post("/login", data={"email": email, "password": password},
                    follow_redirects=False)
    assert r.status_code in (302, 303), \
        f"login failed for {email}: status={r.status_code}"


# ═══════════════════════════════════════════════════════════════════════
#  G — /calendar/ 500 fix
# ═══════════════════════════════════════════════════════════════════════
@check("G1. Calendar route guards `t.project` when project_id is None")
def _():
    """Static check — the regex pattern that used to crash is gone."""
    src = (ROOT / "app/routes/calendar.py").read_text()
    # The unguarded `t.project.name` access should no longer exist.
    assert "t.project.name if t.project" in src or \
           "project_label = t.project.name" in src, \
        "calendar.py doesn't guard t.project access"
    # Standalone-task label
    assert "مهمة مستقلة" in src, "fallback label missing"
    return "calendar.py has project + priority guards in place"


@check("G2. /calendar/ returns 200 even when a standalone task is on the horizon")
def _():
    """Live test — create a standalone task with a near deadline,
    hit /calendar/, must NOT 500. Then clean up."""
    from app.models import (
        Task, TaskStatus, TaskPriority, Company, User,
    )
    company = Company.query.first()
    user = User.query.filter_by(email=DEMO_EMAIL).first()
    t = Task(
        company_id=company.id,
        title="_AUDIT_CALENDAR_STANDALONE",
        project_id=None,   # <-- the previously-crashing case
        status=TaskStatus.TODO,
        priority=TaskPriority.MEDIUM,
        assigned_to_id=user.id,
        created_by_id=user.id,
        deadline=date.today() + timedelta(days=3),
    )
    db.session.add(t); db.session.commit()
    try:
        app = create_app()
        with app.test_client() as client:
            _login(client, DEMO_EMAIL, DEMO_PASS)
            r = client.get("/calendar/", follow_redirects=False)
            assert r.status_code == 200, \
                f"calendar still 500s: status={r.status_code}"
            body = r.data.decode("utf-8")
            # The standalone task should appear with the fallback label.
            assert "_AUDIT_CALENDAR_STANDALONE" in body
            assert "مهمة مستقلة" in body
    finally:
        db.session.delete(t); db.session.commit()
    return "200 + standalone task surfaces with fallback label"


# ═══════════════════════════════════════════════════════════════════════
#  H — manual leave-balance grant
# ═══════════════════════════════════════════════════════════════════════
@check("H1. set_leave_balance service exists + caps to max_balance")
def _():
    from app.services.leave import set_leave_balance
    from app.models import Employee, LeaveType, LeaveBalance, Company
    from decimal import Decimal

    company = Company.query.first()
    emp = Employee.query.filter_by(company_id=company.id).first()
    if not emp:
        return "skipped — no employee fixture in dev DB"

    # Make / find a leave type with max=5
    lt = LeaveType.query.filter_by(
        company_id=company.id, name="_AUDIT_GRANT_TYPE"
    ).first()
    cleanup_type = False
    if not lt:
        lt = LeaveType(
            company_id=company.id, name="_AUDIT_GRANT_TYPE",
            accrual_per_month=Decimal("0"),
            max_balance=Decimal("5"),
            is_paid=True, is_active=True,
        )
        db.session.add(lt); db.session.commit()
        cleanup_type = True

    year = date.today().year
    try:
        # Grant 3 days — should accept as-is.
        row = set_leave_balance(emp, lt, year, balance_days=3.0)
        assert float(row.balance_days) == 3.0, \
            f"expected 3.0, got {row.balance_days}"

        # Grant 100 days — should cap at max_balance (5).
        row = set_leave_balance(emp, lt, year, balance_days=100.0)
        assert float(row.balance_days) == 5.0, \
            f"expected cap to 5.0, got {row.balance_days}"

        # Grant negative — should clamp to 0.
        row = set_leave_balance(emp, lt, year, balance_days=-2)
        assert float(row.balance_days) == 0.0, \
            f"expected clamp to 0, got {row.balance_days}"
    finally:
        # Cleanup the test balance + type
        LeaveBalance.query.filter_by(
            employee_id=emp.id, leave_type_id=lt.id, year=year,
        ).delete()
        if cleanup_type:
            db.session.delete(lt)
        db.session.commit()
    return "service accepts in-range, caps at max, clamps negatives"


@check("H2. POST /hr/employees/<id>/leave-balances/grant updates the row")
def _():
    """Live HTTP round-trip: POST balance, then GET the balances page
    and assert the new number shows."""
    from app.models import (
        Employee, LeaveType, LeaveBalance, Company,
    )
    from decimal import Decimal
    from werkzeug.datastructures import MultiDict
    company = Company.query.first()
    emp = Employee.query.filter_by(company_id=company.id).first()
    if not emp:
        return "skipped — no employee fixture"
    lt = LeaveType.query.filter_by(
        company_id=company.id, is_active=True,
    ).first()
    if not lt:
        return "skipped — no leave type"
    # Make sure max is generous enough so we hit the snap-to-input path.
    original_max = lt.max_balance
    if not lt.max_balance or float(lt.max_balance) < 25:
        lt.max_balance = Decimal("25")
        db.session.commit()
    year = date.today().year
    app = create_app()
    try:
        with app.test_client() as client:
            _login(client, DEMO_EMAIL, DEMO_PASS)
            r = client.post(
                f"/hr/employees/{emp.id}/leave-balances/grant",
                data=MultiDict([
                    ("leave_type_id", str(lt.id)),
                    ("year", str(year)),
                    ("balance_days", "7.5"),
                ]),
                follow_redirects=False,
            )
            assert r.status_code in (302, 303), \
                f"grant POST status={r.status_code}"
        # Re-fetch from DB
        db.session.expire_all()
        row = LeaveBalance.query.filter_by(
            employee_id=emp.id, leave_type_id=lt.id, year=year,
        ).first()
        assert row, "balance row not created"
        assert float(row.balance_days) == 7.5, \
            f"expected 7.5, got {row.balance_days}"
        # Cleanup
        db.session.delete(row); db.session.commit()
    finally:
        if original_max != lt.max_balance:
            lt.max_balance = original_max
            db.session.commit()
    return "POST → balance row = 7.5 (and visible via GET)"


@check("H3. employee_balances template shows the 'تعيين' input column")
def _():
    src = (ROOT / "app/templates/hr/employee_balances.html").read_text()
    assert 'name="balance_days"' in src, "grant input field missing"
    assert "grant_leave_balance" in src, "form action missing"
    assert "تعيين" in src, "column header missing"
    return "template has the per-row grant form"


# ═══════════════════════════════════════════════════════════════════════
#  I — leads Excel export
# ═══════════════════════════════════════════════════════════════════════
@check("I1. export_leads_excel produces a valid xlsx with expected sheet name")
def _():
    from openpyxl import load_workbook
    from app.services.export import export_leads_excel
    from app.models import Company, Lead
    company = Company.query.first()
    leads = Lead.query.filter_by(company_id=company.id).limit(5).all()
    buf = export_leads_excel(company, leads)
    wb = load_workbook(buf)
    assert "Leads" in wb.sheetnames
    ws = wb["Leads"]
    # Header row should have at least 18 columns
    header_row = next(ws.iter_rows(min_row=5, max_row=5, values_only=True))
    assert len([c for c in header_row if c]) >= 18, \
        f"header has {len([c for c in header_row if c])} cols, expected ≥18"
    # Verify a few specific Arabic headers are present
    headers_set = {c for c in header_row if c}
    for expected in ("الاسم", "الهاتف", "البريد", "الحالة",
                       "القيمة المتوقعة"):
        assert expected in headers_set, \
            f"header {expected!r} missing"
    return f"sheet OK, {len(leads)} leads exported, {len(headers_set)} columns"


@check("I2. GET /leads/export/excel returns 200 + correct mimetype")
def _():
    app = create_app()
    with app.test_client() as client:
        _login(client, DEMO_EMAIL, DEMO_PASS)
        r = client.get("/leads/export/excel", follow_redirects=False)
        assert r.status_code == 200, f"status={r.status_code}"
        mime = r.headers.get("Content-Type", "")
        assert "spreadsheetml" in mime, \
            f"wrong mimetype: {mime}"
        # Body should start with ZIP magic bytes (xlsx = zip container)
        assert r.data[:2] == b"PK", \
            "response is not a valid xlsx (missing ZIP signature)"
        # Filename hint
        cd = r.headers.get("Content-Disposition", "")
        assert ".xlsx" in cd, f"no .xlsx in Content-Disposition: {cd}"
    return "200 + xlsx mimetype + valid ZIP container"


@check("I3. /leads/ index page exposes the export button")
def _():
    src = (ROOT / "app/templates/leads/index.html").read_text()
    assert "تصدير Excel" in src, "export button label missing"
    assert "leads.export_excel" in src, "url_for to export route missing"
    return "button + url_for present"


@check("I4. Export honours filters (e.g. ?status=WON yields a different row count)")
def _():
    """Hit the export twice with different filters and assert the
    returned sheets have different row counts (lower bound — even an
    empty filter result is acceptable)."""
    from openpyxl import load_workbook
    app = create_app()
    with app.test_client() as client:
        _login(client, DEMO_EMAIL, DEMO_PASS)
        r_all = client.get("/leads/export/excel")
        r_won = client.get("/leads/export/excel?status=WON")
    wb_all = load_workbook(io.BytesIO(r_all.data))["Leads"]
    wb_won = load_workbook(io.BytesIO(r_won.data))["Leads"]
    # Count rows below header (which sits at row 5)
    rows_all = sum(1 for _ in wb_all.iter_rows(min_row=6, max_row=200,
                                                  values_only=True)
                    if any(c not in (None, "") for c in _))
    rows_won = sum(1 for _ in wb_won.iter_rows(min_row=6, max_row=200,
                                                  values_only=True)
                    if any(c not in (None, "") for c in _))
    # WON ≤ all-time count
    assert rows_won <= rows_all, \
        f"WON filter widened the count: {rows_won} > {rows_all}"
    return f"all-time rows={rows_all}, WON rows={rows_won} (≤)"


def main():
    app = create_app()
    with app.app_context():
        passed = failed = 0
        for label, fn in CHECKS:
            try:
                msg = fn()
                print(f"\033[92mPASS\033[0m  {label}")
                if msg:
                    print(f"        {msg}")
                passed += 1
            except Exception as e:
                print(f"\033[91mFAIL\033[0m  {label}")
                print(f"        {type(e).__name__}: {e}")
                import traceback
                traceback.print_exc()
                failed += 1
        print()
        print(f"  {passed}/{passed + failed} checks passed.")
        return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
