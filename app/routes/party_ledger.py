"""MARSOUD-PARTY-LEDGER-02 — UI for the unified party statement."""
import io
from datetime import date, datetime, timedelta

from flask import (
    Blueprint, render_template, request, g, send_file, flash, redirect, url_for,
)
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.services.party_ledger import (
    list_parties, party_ledger, KIND_LABELS,
)
from app.services.permissions import require_permission


bp = Blueprint("party_ledger", __name__)


def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _default_window():
    today = date.today()
    return today - timedelta(days=90), today


@bp.route("/", methods=["GET"])
@login_required
@require_permission("reports.view")
def index():
    cid = g.active_company.id
    kind = (request.args.get("kind") or "customer").lower()
    if kind not in KIND_LABELS:
        kind = "customer"
    party_id = request.args.get("party_id", type=int)
    start = _parse_date(request.args.get("start_date"))
    end = _parse_date(request.args.get("end_date"))
    if not start and not end:
        start, end = _default_window()
    elif not end:
        end = date.today()

    parties = list_parties(cid, kind)
    statement = None
    if party_id:
        try:
            statement = party_ledger(cid, kind, party_id,
                                       start_date=start, end_date=end)
        except ValueError as e:
            flash(str(e), "error")
    return render_template(
        "party_ledger/index.html",
        kind=kind, kind_label=KIND_LABELS[kind],
        kind_labels=KIND_LABELS,
        parties=parties, party_id=party_id,
        start_date=start, end_date=end,
        statement=statement,
    )


@bp.route("/export.pdf", methods=["GET"])
@login_required
@require_permission("reports.export")
def export_pdf():
    """Render the statement to a printable PDF (uses the existing
    Playwright Chromium that's already a dev dependency)."""
    from playwright.sync_api import sync_playwright
    import tempfile
    cid = g.active_company.id
    kind = (request.args.get("kind") or "customer").lower()
    party_id = request.args.get("party_id", type=int)
    if kind not in KIND_LABELS or not party_id:
        flash("اختر الطرف الأول", "error")
        return redirect(url_for("party_ledger.index", kind=kind))
    start = _parse_date(request.args.get("start_date"))
    end = _parse_date(request.args.get("end_date"))
    if not start and not end:
        start, end = _default_window()
    elif not end:
        end = date.today()

    statement = party_ledger(cid, kind, party_id,
                              start_date=start, end_date=end)

    html = render_template(
        "party_ledger/print.html",
        statement=statement, company=g.active_company,
        start_date=start, end_date=end,
    )
    # Render via headless Chromium → PDF bytes
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
        f.write(html.encode("utf-8"))
        html_path = f.name
    pdf_buf = io.BytesIO()
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(f"file://{html_path}", wait_until="networkidle")
            pdf_bytes = page.pdf(
                format="A4",
                margin={"top": "1.5cm", "bottom": "1.5cm",
                         "left": "1cm", "right": "1cm"},
                print_background=True,
            )
            pdf_buf.write(pdf_bytes)
            browser.close()
    finally:
        import os
        os.unlink(html_path)
    pdf_buf.seek(0)

    safe_name = "".join(ch if ch.isalnum() else "_"
                         for ch in statement["party"]["name"])[:40]
    fn = f"ledger-{safe_name}-{date.today().isoformat()}.pdf"
    return send_file(pdf_buf, mimetype="application/pdf",
                     as_attachment=True, download_name=fn)


@bp.route("/export.xlsx", methods=["GET"])
@login_required
@require_permission("reports.export")
def export_xlsx():
    cid = g.active_company.id
    kind = (request.args.get("kind") or "customer").lower()
    party_id = request.args.get("party_id", type=int)
    if kind not in KIND_LABELS or not party_id:
        flash("اختر الطرف الأول", "error")
        return redirect(url_for("party_ledger.index", kind=kind))
    start = _parse_date(request.args.get("start_date"))
    end = _parse_date(request.args.get("end_date"))
    if not start and not end:
        start, end = _default_window()
    elif not end:
        end = date.today()

    statement = party_ledger(cid, kind, party_id,
                              start_date=start, end_date=end)

    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الحساب"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A2540")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    label_font = Font(bold=True)

    # Title block
    ws.cell(row=1, column=1, value=f"كشف حساب — {statement['party']['name']}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"النوع: {statement['party']['kind_label']}  "
                                     f"·  الحساب: {statement['party']['account_code']}")
    if start or end:
        rng = f"من {start.isoformat() if start else '...'} إلى {end.isoformat() if end else '...'}"
        ws.cell(row=3, column=1, value=rng)

    # Opening balance row
    ws.cell(row=5, column=1, value="الرصيد الافتتاحي").font = label_font
    ws.cell(row=5, column=6, value=statement["opening_balance"]).number_format = "#,##0.00"

    # Header row
    headers = ["التاريخ", "رقم القيد", "البيان", "المرجع", "ملاحظة",
                "مدين", "دائن", "الرصيد"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=7, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    r = 8
    for row in statement["rows"]:
        ws.cell(row=r, column=1, value=row["date"].isoformat() if row["date"] else "")
        ws.cell(row=r, column=2, value=row["entry_number"] or "")
        ws.cell(row=r, column=3, value=row["description"] or "")
        ws.cell(row=r, column=4, value=row["source_type"] or "")
        ws.cell(row=r, column=5, value=row["memo"] or "")
        ws.cell(row=r, column=6, value=row["debit"] or 0).number_format = "#,##0.00"
        ws.cell(row=r, column=7, value=row["credit"] or 0).number_format = "#,##0.00"
        ws.cell(row=r, column=8, value=row["balance"] or 0).number_format = "#,##0.00"
        r += 1

    # Totals
    ws.cell(row=r + 1, column=1, value="الإجماليات").font = label_font
    ws.cell(row=r + 1, column=6, value=statement["total_debit"]).number_format = "#,##0.00"
    ws.cell(row=r + 1, column=6).font = label_font
    ws.cell(row=r + 1, column=7, value=statement["total_credit"]).number_format = "#,##0.00"
    ws.cell(row=r + 1, column=7).font = label_font
    ws.cell(row=r + 1, column=8, value=statement["closing_balance"]).number_format = "#,##0.00"
    ws.cell(row=r + 1, column=8).font = label_font

    # Column widths
    widths = [12, 14, 40, 18, 30, 14, 14, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w
    ws.freeze_panes = "A8"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(ch if ch.isalnum() else "_"
                         for ch in statement["party"]["name"])[:40]
    fn = f"ledger-{safe_name}-{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fn,
    )
