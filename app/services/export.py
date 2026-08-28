"""PDF and Excel export for financial reports."""
import io
import os
import base64
import logging
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from flask import render_template
from app import db
# MARSOUD-CURRENCY-AR — the ReportLab renderers build strings in Python,
# so they can't use the `currency_ar` Jinja filter.
from app.services.currency import currency_name_ar

_export_logger = logging.getLogger("ledgeros.export")
from bidi.algorithm import get_display
from app.services.reports import (
    balance_sheet, income_statement, cash_flow,
    income_summary, expenses_summary, income_statement_compared,
    aging_report, ap_aging_report, vat_report,
    payroll_summary_report, fixed_assets_report,
)

NAVY = colors.HexColor("#0A2540")
BLUE = colors.HexColor("#2563EB")
GRAY = colors.HexColor("#64748B")

# ─── Arabic-capable font registration ──────────────────────────────────
_FONT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "fonts")
_FONT_REGULAR = "Amiri"
_FONT_BOLD = "Amiri-Bold"


def _register_fonts():
    if _FONT_REGULAR in pdfmetrics.getRegisteredFontNames():
        return
    pdfmetrics.registerFont(TTFont(_FONT_REGULAR, os.path.join(_FONT_DIR, "Amiri-Regular.ttf")))
    pdfmetrics.registerFont(TTFont(_FONT_BOLD, os.path.join(_FONT_DIR, "Amiri-Bold.ttf")))


_register_fonts()


def ar(text):
    """Shape Arabic text for correct rendering in reportlab PDFs.

    Safe to call on any string: reshapes Arabic ligatures + applies bidi
    for correct visual order. Pure-Latin strings pass through unchanged.
    """
    if text is None:
        return ""
    s = str(text)
    if not s:
        return s
    reshaped = arabic_reshaper.reshape(s)
    return get_display(reshaped)


def _excel_styled_header(ws, title, company_name, period):
    ws["A1"] = company_name
    ws["A1"].font = Font(size=16, bold=True, color="0A2540")
    ws["A2"] = title
    ws["A2"].font = Font(size=14, bold=True, color="2563EB")
    ws["A3"] = period
    ws["A3"].font = Font(size=10, color="64748B")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def export_balance_sheet_excel(company, as_of):
    data = balance_sheet(company.id, as_of=as_of)
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    _excel_styled_header(ws, "Balance Sheet — الميزانية العمومية", company.name, f"كما في {as_of}")

    row = 5
    ws.cell(row=row, column=1, value="الأصول").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["assets"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي الأصول").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["totals"]["assets"]).font = Font(bold=True)
    ws.cell(row=row, column=2).number_format = "#,##0.00"
    row += 2

    ws.cell(row=row, column=1, value="الالتزامات").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["liabilities"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي الالتزامات").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["totals"]["liabilities"]).font = Font(bold=True)
    row += 2

    ws.cell(row=row, column=1, value="حقوق الملكية").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["equity"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي حقوق الملكية").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["totals"]["equity"]).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_income_statement_excel(company, start, end):
    data = income_statement(company.id, start_date=start, end_date=end)
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    _excel_styled_header(ws, "Income Statement — قائمة الدخل", company.name, f"من {start} إلى {end}")

    row = 5
    ws.cell(row=row, column=1, value="الإيرادات").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["revenue"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي الإيرادات").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["total_revenue"]).font = Font(bold=True)
    row += 2

    ws.cell(row=row, column=1, value="المصروفات").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for item in data["expenses"]:
        ws.cell(row=row, column=1, value=f"  {item['code']} — {item['name']}")
        ws.cell(row=row, column=2, value=item["balance"]).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="إجمالي المصروفات").font = Font(bold=True)
    ws.cell(row=row, column=2, value=data["total_expense"]).font = Font(bold=True)
    row += 2

    color = "10B981" if data["net_income"] >= 0 else "EF4444"
    ws.cell(row=row, column=1, value="صافي الربح / الخسارة").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=color)
    ws.cell(row=row, column=2, value=data["net_income"]).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=color)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── MARSOUD-51: WeasyPrint helpers for the new HTML→PDF flow ─────────
def _company_logo_data_uri(company):
    """Return the company logo as a data: URI so WeasyPrint embeds it inline.
    None if the company has no logo or the file is missing."""
    if not getattr(company, "logo_path", None):
        return None
    rel = company.logo_path.lstrip("/")
    candidate = os.path.join(os.path.dirname(os.path.dirname(__file__)), rel)
    if not os.path.exists(candidate):
        return None
    ext = os.path.splitext(candidate)[1].lstrip(".").lower()
    mime = "image/png" if ext == "png" else f"image/{ext}"
    with open(candidate, "rb") as f:
        return f"data:{mime};base64,{base64.b64encode(f.read()).decode('ascii')}"


# ─── MARSOUD-PDF-P0 (2026-08-28) — Amiri font embedded as data: URIs ──
# The three PDF templates (pdfs/invoice, pdfs/payslip, party_ledger/print)
# used to declare `font-family: 'Tajawal', 'Cairo', ...` — but neither TTF
# is on disk. WeasyPrint (and Chromium, for party_ledger) silently fell back
# to whatever Arabic font the deploy host had, so identical invoices from
# two hosts looked visibly different. Only Amiri-Regular.ttf +
# Amiri-Bold.ttf are actually on disk in app/static/fonts/, so we embed
# them as data: URIs. Data-URI form works for both engines identically:
#  - WeasyPrint doesn't need base_url resolution.
#  - Chromium loading a file:// tempfile from the OS temp dir can't
#    resolve relative url() → the tempfile isn't next to app/static/,
#    so a data URI is the only portable option.
# The helper is cached at module scope — the base64 encode is ~200KB per
# weight, done once per process. If the TTFs are missing (fresh clone
# without the fonts) we return an empty <style/> block so the app still
# boots; the PDF then renders in the OS default, which is what happens
# today anyway (i.e. this change strictly cannot make things worse).
_AMIRI_FONT_FACE_CSS_CACHE = None


def _amiri_font_face_css():
    """Return `<style>@font-face …</style>` for Amiri Regular + Bold with the
    TTFs inlined as data: URIs. Templates include the returned block
    verbatim (via `{{ amiri_font_face|safe }}`) OUTSIDE their own <style>
    tag — the browser/WeasyPrint concatenates the two stylesheets normally,
    and we avoid the invalid "<style> inside <style>" nesting."""
    global _AMIRI_FONT_FACE_CSS_CACHE
    if _AMIRI_FONT_FACE_CSS_CACHE is not None:
        return _AMIRI_FONT_FACE_CSS_CACHE
    fonts_dir = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "static", "fonts")
    faces = []
    for weight, filename in (
        (400, "Amiri-Regular.ttf"),
        (700, "Amiri-Bold.ttf"),
    ):
        path = os.path.join(fonts_dir, filename)
        if not os.path.exists(path):
            continue
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")
        faces.append(
            "@font-face { font-family: 'Amiri'; "
            f"src: url(data:font/ttf;base64,{b64}) format('truetype'); "
            f"font-weight: {weight}; font-style: normal; }}"
        )
    _AMIRI_FONT_FACE_CSS_CACHE = "<style>\n" + "\n".join(faces) + "\n</style>"
    return _AMIRI_FONT_FACE_CSS_CACHE


def _weasyprint_render(template_name, **context):
    """Render a Jinja template + run it through WeasyPrint. Returns a BytesIO
    with the PDF. WeasyPrint is imported lazily so the rest of the app keeps
    booting on a host that doesn't have libpango installed."""
    import weasyprint  # lazy — system libs (libpango/cairo) only needed for PDFs
    # MARSOUD-PDF-P0 — every WeasyPrint template that needs Arabic type
    # references {{ amiri_font_face|safe }}; inject it here so callers
    # don't each have to remember. Callers may override by passing their
    # own `amiri_font_face` in **context.
    context.setdefault("amiri_font_face", _amiri_font_face_css())
    html = render_template(template_name, **context)
    buf = io.BytesIO()
    weasyprint.HTML(string=html, base_url=os.path.dirname(os.path.dirname(__file__))).write_pdf(buf)
    buf.seek(0)
    return buf


def _company_logo_disk_path(company):
    """Resolve company.logo_path (a `/static/...` URL) to an absolute disk path
    suitable for reportlab. Returns None if the file doesn't exist locally
    (e.g. logo_url is an external URL not on disk).
    """
    if not getattr(company, "logo_path", None):
        return None
    rel = company.logo_path.lstrip("/")
    # Static lives at app/static/ — same as the export module's parent's static.
    candidate = os.path.join(os.path.dirname(os.path.dirname(__file__)), rel)
    return candidate if os.path.exists(candidate) else None


def _pdf_header(p, company, title, period):
    p.setFillColor(NAVY)
    p.rect(0, 27.7 * cm, 21 * cm, 2 * cm, fill=1, stroke=0)

    # MARSOUD-23 — draw company logo on the left when present
    logo_x = 1.5 * cm
    logo_disk = _company_logo_disk_path(company)
    if logo_disk:
        try:
            from reportlab.lib.utils import ImageReader
            img = ImageReader(logo_disk)
            # Logo box: 2cm tall, max 3cm wide, anchored on the navy band
            p.drawImage(img, 1.2 * cm, 27.9 * cm, width=3 * cm, height=1.6 * cm,
                        preserveAspectRatio=True, mask="auto")
            logo_x = 4.7 * cm
        except Exception:
            pass

    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 18)
    p.drawString(logo_x, 28.5 * cm, ar(company.name))
    p.setFont(_FONT_REGULAR, 10)
    sub = "Marsoud — Financial Report"
    if getattr(company, "tax_number", None):
        sub += f"  ·  VAT # {company.tax_number}"
    p.drawString(logo_x, 28 * cm, ar(sub))

    p.setFillColor(NAVY)
    p.setFont(_FONT_BOLD, 16)
    p.drawString(1.5 * cm, 26.5 * cm, ar(title))
    p.setFillColor(GRAY)
    p.setFont(_FONT_REGULAR, 10)
    p.drawString(1.5 * cm, 26 * cm, ar(period))


def _pdf_section(p, y, label):
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 11)
    p.drawString(1.3 * cm, y - 0.2 * cm, ar(label))
    return y - 1 * cm


def export_balance_sheet_pdf(company, as_of):
    data = balance_sheet(company.id, as_of=as_of)
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, company, "Balance Sheet", f"As of {as_of}")

    y = 24.5 * cm
    for section, items, total_key in [
        ("ASSETS", data["assets"], "assets"),
        ("LIABILITIES", data["liabilities"], "liabilities"),
        ("EQUITY", data["equity"], "equity"),
    ]:
        y = _pdf_section(p, y, section)
        p.setFillColor(colors.black)
        p.setFont(_FONT_REGULAR, 10)
        for item in items:
            if y < 3 * cm:
                p.showPage()
                _pdf_header(p, company, "Balance Sheet (cont.)", f"As of {as_of}")
                y = 24.5 * cm
            p.drawString(1.5 * cm, y, ar(f"{item['code']}  {item['name']}"))
            p.drawRightString(19.5 * cm, y, f"{item['balance']:,.2f}")
            y -= 0.5 * cm
        p.setFont(_FONT_BOLD, 10)
        p.setFillColor(BLUE)
        p.drawString(1.5 * cm, y, ar(f"Total {section}"))
        p.drawRightString(19.5 * cm, y, f"{data['totals'][total_key]:,.2f}")
        p.setFillColor(colors.black)
        y -= 1 * cm

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_income_statement_pdf(company, start, end):
    data = income_statement(company.id, start_date=start, end_date=end)
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, company, "Income Statement", f"{start} → {end}")

    y = 24.5 * cm
    for section, items, total_key in [
        ("REVENUE", data["revenue"], "total_revenue"),
        ("EXPENSES", data["expenses"], "total_expense"),
    ]:
        y = _pdf_section(p, y, section)
        p.setFillColor(colors.black)
        p.setFont(_FONT_REGULAR, 10)
        for item in items:
            p.drawString(1.5 * cm, y, ar(f"{item['code']}  {item['name']}"))
            p.drawRightString(19.5 * cm, y, f"{item['balance']:,.2f}")
            y -= 0.5 * cm
        p.setFont(_FONT_BOLD, 10)
        p.setFillColor(BLUE)
        p.drawString(1.5 * cm, y, ar(f"Total {section}"))
        p.drawRightString(19.5 * cm, y, f"{data[total_key]:,.2f}")
        p.setFillColor(colors.black)
        y -= 1 * cm

    y -= 0.3 * cm
    profit = data["net_income"]
    color = colors.HexColor("#10B981") if profit >= 0 else colors.HexColor("#EF4444")
    p.setFillColor(color)
    p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.9 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 13)
    p.drawString(1.5 * cm, y - 0.2 * cm, "NET PROFIT / (LOSS)")
    p.drawRightString(19.5 * cm, y - 0.2 * cm, f"{profit:,.2f}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


_STATUS_COLOR = {
    "DRAFT":              (colors.HexColor("#475569"), "DRAFT"),
    "SENT":               (colors.HexColor("#2563EB"), "SENT"),
    "PAID":               (colors.HexColor("#10B981"), "PAID"),
    "PARTIALLY_PAID":     (colors.HexColor("#F59E0B"), "PARTIALLY PAID"),
    "OVERDUE":            (colors.HexColor("#EF4444"), "OVERDUE"),
    "CANCELLED":          (colors.HexColor("#64748B"), "CANCELLED"),
    "REFUNDED":           (colors.HexColor("#EC4899"), "REFUNDED"),
    "PARTIALLY_REFUNDED": (colors.HexColor("#EC4899"), "PART. REFUNDED"),
}


def export_invoice_pdf(invoice):
    """MARSOUD-51 — Generate a customer-facing invoice PDF via WeasyPrint
    using the new branded HTML template. Falls back to the legacy ReportLab
    renderer if WeasyPrint fails to load (e.g. missing libpango on host)."""
    try:
        return _weasyprint_render(
            "pdfs/invoice.html",
            invoice=invoice,
            company_logo_data_uri=_company_logo_data_uri(invoice.company),
        )
    except Exception as e:
        # MARSOUD-54.1 — make the silent regression loud. Old log line was
        # logger.exception() which is buried inside a normal log feed; this
        # prints a clearly tagged warning so deploy issues are obvious.
        _export_logger.warning(
            "[PDF-FALLBACK] Invoice %s: WeasyPrint failed (%s: %s) — "
            "rendering the OLD ReportLab layout instead. This means the "
            "customer is NOT getting the new branded PDF. Cause is almost "
            "always missing weasyprint pip install OR missing libpango on "
            "the host. Run: pip install -r requirements.txt + "
            "apt install libpango-1.0-0 libpangoft2-1.0-0 libcairo2",
            invoice.number, type(e).__name__, str(e)[:200],
        )
        return _export_invoice_pdf_legacy(invoice)


def _export_invoice_pdf_legacy(invoice):
    """Legacy ReportLab-based invoice PDF kept as fallback."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, invoice.company,
                f"Invoice {invoice.number}",
                f"Issue: {invoice.issue_date}  ·  Due: {invoice.due_date}")

    # ─── Status pill in top-left of header band ─────────────────────────
    status_value = invoice.status.value if hasattr(invoice.status, "value") else str(invoice.status)
    badge_color, badge_label = _STATUS_COLOR.get(status_value, (NAVY, status_value))
    p.setFillColor(badge_color)
    p.rect(15.5 * cm, 28.4 * cm, 4.5 * cm, 0.9 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 11)
    p.drawCentredString(15.5 * cm + 2.25 * cm, 28.7 * cm, badge_label)

    y = 24.5 * cm

    # ─── Customer + dates two-column cards ──────────────────────────────
    card_h = 2.2 * cm
    # Left: BILL TO
    p.setFillColor(colors.HexColor("#F8FAFC"))
    p.roundRect(1 * cm, y - card_h, 9 * cm, card_h, 6, fill=1, stroke=0)
    p.setFillColor(colors.HexColor("#94A3B8"))
    p.setFont(_FONT_BOLD, 9)
    p.drawString(1.4 * cm, y - 0.55 * cm, "BILL TO")
    p.setFillColor(colors.HexColor("#0A2540"))
    p.setFont(_FONT_BOLD, 12)
    p.drawString(1.4 * cm, y - 1.1 * cm, ar(invoice.customer.name))
    p.setFont(_FONT_REGULAR, 9)
    p.setFillColor(colors.HexColor("#475569"))
    line_y = y - 1.55 * cm
    if invoice.customer.email:
        p.drawString(1.4 * cm, line_y, invoice.customer.email)
        line_y -= 0.4 * cm
    if getattr(invoice.customer, "phone", None):
        p.drawString(1.4 * cm, line_y, invoice.customer.phone)

    # Right: DATES + CURRENCY
    p.setFillColor(colors.HexColor("#F8FAFC"))
    p.roundRect(11 * cm, y - card_h, 9 * cm, card_h, 6, fill=1, stroke=0)
    p.setFillColor(colors.HexColor("#94A3B8"))
    p.setFont(_FONT_BOLD, 9)
    p.drawString(11.4 * cm, y - 0.55 * cm, "INVOICE DETAILS")
    p.setFillColor(colors.HexColor("#0F172A"))
    p.setFont(_FONT_REGULAR, 9)
    p.drawString(11.4 * cm, y - 1.1 * cm, f"Issued: {invoice.issue_date.isoformat()}")
    p.drawString(11.4 * cm, y - 1.5 * cm, f"Due:    {invoice.due_date.isoformat()}")
    p.drawString(11.4 * cm, y - 1.9 * cm, f"Currency: {currency_name_ar(invoice.currency)}")

    y -= card_h + 0.8 * cm

    # ─── Items table header ─────────────────────────────────────────────
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 10)
    p.drawString(1.3 * cm, y - 0.15 * cm, "DESCRIPTION")
    p.drawString(12 * cm, y - 0.15 * cm, "QTY")
    p.drawString(14 * cm, y - 0.15 * cm, "PRICE")
    p.drawRightString(19.5 * cm, y - 0.15 * cm, "TOTAL")
    y -= 0.9 * cm

    # ─── Items rows with alternating row shading ───────────────────────
    p.setFont(_FONT_REGULAR, 10)
    row_idx = 0
    for item in invoice.items:
        if y < 6 * cm:
            p.showPage()
            _pdf_header(p, invoice.company,
                        f"Invoice {invoice.number} (cont.)",
                        f"Issue: {invoice.issue_date}  ·  Due: {invoice.due_date}")
            y = 24.5 * cm
        if row_idx % 2 == 1:
            p.setFillColor(colors.HexColor("#F8FAFC"))
            p.rect(1 * cm, y - 0.25 * cm, 19 * cm, 0.6 * cm, fill=1, stroke=0)
        p.setFillColor(colors.black)
        p.drawString(1.3 * cm, y, ar(item.description[:60]))
        p.drawString(12 * cm, y, f"{float(item.quantity):.2f}")
        p.drawString(14 * cm, y, f"{float(item.unit_price):,.2f}")
        p.drawRightString(19.5 * cm, y, f"{float(item.line_total):,.2f}")
        y -= 0.55 * cm
        row_idx += 1

    # ─── Totals card (right-aligned, fixed-width) ──────────────────────
    totals_x = 11 * cm
    totals_w = 9 * cm
    y -= 0.5 * cm
    rows = [
        ("Subtotal",                  f"{float(invoice.subtotal):,.2f}",          False),
    ]
    if getattr(invoice, "invoice_discount_amount", 0) and float(invoice.invoice_discount_amount) > 0:
        rows.append(("Discount", f"-{float(invoice.invoice_discount_amount):,.2f}", False))
    rows.append((f"VAT ({float(invoice.tax_rate):.0f}%)", f"{float(invoice.tax_amount):,.2f}", False))
    rows.append(("Total",     f"{float(invoice.total):,.2f} {currency_name_ar(invoice.currency)}",  True))
    if invoice.paid_amount and float(invoice.paid_amount) > 0:
        rows.append(("Paid",    f"{float(invoice.paid_amount):,.2f}",              False))
        rows.append(("Balance Due", f"{float(invoice.balance):,.2f} {currency_name_ar(invoice.currency)}", True))

    card_h = 0.7 * cm * len(rows) + 0.3 * cm
    p.setFillColor(colors.HexColor("#F8FAFC"))
    p.roundRect(totals_x, y - card_h, totals_w, card_h, 6, fill=1, stroke=0)

    inner_y = y - 0.5 * cm
    for label, value, emphasize in rows:
        if emphasize:
            p.setFillColor(NAVY)
            p.rect(totals_x + 0.1 * cm, inner_y - 0.18 * cm, totals_w - 0.2 * cm, 0.55 * cm, fill=1, stroke=0)
            p.setFillColor(colors.white)
            p.setFont(_FONT_BOLD, 11)
        else:
            p.setFillColor(colors.HexColor("#475569"))
            p.setFont(_FONT_REGULAR, 10)
        p.drawString(totals_x + 0.4 * cm, inner_y, label)
        p.drawRightString(totals_x + totals_w - 0.4 * cm, inner_y, value)
        inner_y -= 0.65 * cm

    # ─── Footer ─────────────────────────────────────────────────────────
    p.setFillColor(colors.HexColor("#94A3B8"))
    p.setFont(_FONT_REGULAR, 8)
    footer = f"Thank you for your business — {invoice.company.name}"
    if invoice.company.tax_number:
        footer += f"  ·  VAT # {invoice.company.tax_number}"
    p.drawCentredString(10.5 * cm, 1.2 * cm, ar(footer))

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_payslip_pdf(employee, line, run):
    """MARSOUD-51 — Generate a payslip PDF via WeasyPrint using the new
    branded HTML template. Falls back to the legacy ReportLab renderer if
    WeasyPrint fails to load."""
    try:
        return _weasyprint_render(
            "pdfs/payslip.html",
            employee=employee, line=line, run=run,
            company_logo_data_uri=_company_logo_data_uri(employee.company),
            absences_count=int(line.absences_count or 0),
            overtime_hours=float(line.overtime_hours or 0),
            leaves_count=0,
            late_hours=0,
        )
    except Exception as e:
        # MARSOUD-54.1 — loud warning so the silent regression is visible.
        _export_logger.warning(
            "[PDF-FALLBACK] Payslip %s/%s for %s: WeasyPrint failed "
            "(%s: %s) — rendering the OLD ReportLab layout instead. Run "
            "pip install -r requirements.txt + apt install "
            "libpango-1.0-0 libpangoft2-1.0-0 libcairo2 on the host.",
            run.period_month, run.period_year, employee.name,
            type(e).__name__, str(e)[:200],
        )
        return _export_payslip_pdf_legacy(employee, line, run)


def _export_payslip_pdf_legacy(employee, line, run):
    """Legacy ReportLab-based payslip PDF kept as fallback."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(
        p, employee.company,
        f"Payslip — {run.period_month}/{run.period_year}",
        f"Payroll Run {run.number or run.id}",
    )

    y = 24.5 * cm

    # ─── Employee identity card ────────────────────────────────────────
    card_h = 2.4 * cm
    p.setFillColor(colors.HexColor("#F8FAFC"))
    p.roundRect(1 * cm, y - card_h, 19 * cm, card_h, 6, fill=1, stroke=0)
    p.setFillColor(colors.HexColor("#94A3B8"))
    p.setFont(_FONT_BOLD, 9)
    p.drawString(1.4 * cm, y - 0.55 * cm, "EMPLOYEE")
    p.setFillColor(colors.HexColor("#0A2540"))
    p.setFont(_FONT_BOLD, 14)
    p.drawString(1.4 * cm, y - 1.15 * cm, ar(employee.name))
    p.setFont(_FONT_REGULAR, 10)
    p.setFillColor(colors.HexColor("#475569"))
    p.drawString(1.4 * cm, y - 1.65 * cm, ar(employee.job_title or "—"))
    if employee.department:
        p.drawString(1.4 * cm, y - 2.05 * cm, ar(employee.department.name))

    # Right side of the card — metadata column
    meta_x = 12 * cm
    p.setFont(_FONT_REGULAR, 9)
    p.drawString(meta_x, y - 0.7 * cm, f"Employee #: {employee.employee_number or '—'}")
    p.drawString(meta_x, y - 1.1 * cm, f"Period:     {run.period_month:02d}/{run.period_year}")
    p.drawString(meta_x, y - 1.5 * cm, f"Working days: {line.working_days}/30")
    if line.attendance_auto_calculated:
        p.setFillColor(colors.HexColor("#10B981"))
        p.setFont(_FONT_BOLD, 8)
        p.drawString(meta_x, y - 1.9 * cm, "Attendance auto-computed")

    y -= card_h + 0.6 * cm

    # ─── Earnings + deductions side-by-side ────────────────────────────
    earnings = [
        ("Basic (prorated)",  float(line.basic)),
        ("Allowances",        float(line.allowances)),
        ("Overtime",          float(line.overtime)),
        ("Bonus",             float(line.bonus)),
    ]
    deductions = [
        ("Fixed deductions",  float(line.deductions)),
        ("Absence",           float(line.absence_deduction)),
        ("Late",              float(line.late_deduction)),
        ("Advance",           float(line.advance_deduction)),
    ]

    def _section(x, w, title, rows, accent):
        nonlocal_y = y
        # Header strip
        p.setFillColor(accent)
        p.rect(x, nonlocal_y - 0.5 * cm, w, 0.7 * cm, fill=1, stroke=0)
        p.setFillColor(colors.white)
        p.setFont(_FONT_BOLD, 10)
        p.drawString(x + 0.3 * cm, nonlocal_y - 0.3 * cm, title)
        ly = nonlocal_y - 1.1 * cm
        # Card body
        body_h = 0.55 * cm * len(rows) + 0.9 * cm
        p.setFillColor(colors.HexColor("#F8FAFC"))
        p.rect(x, ly - body_h + 0.7 * cm, w, body_h, fill=1, stroke=0)
        p.setFillColor(colors.HexColor("#0F172A"))
        p.setFont(_FONT_REGULAR, 10)
        subtotal = 0.0
        for label, value in rows:
            if abs(value) < 0.005:
                continue
            p.drawString(x + 0.3 * cm, ly, ar(label))
            p.drawRightString(x + w - 0.3 * cm, ly, f"{value:,.2f}")
            subtotal += value
            ly -= 0.55 * cm
        # Subtotal
        p.setFillColor(accent)
        p.setFont(_FONT_BOLD, 10)
        p.drawString(x + 0.3 * cm, ly, "Subtotal")
        p.drawRightString(x + w - 0.3 * cm, ly, f"{subtotal:,.2f}")
        return subtotal, ly - 0.6 * cm

    gross, y_left = _section(1 * cm, 9 * cm, "EARNINGS", earnings,
                             colors.HexColor("#10B981"))
    deduct_total, y_right = _section(11 * cm, 9 * cm, "DEDUCTIONS", deductions,
                                     colors.HexColor("#EF4444"))
    y = min(y_left, y_right) - 0.4 * cm

    # ─── Net Pay banner ─────────────────────────────────────────────────
    p.setFillColor(NAVY)
    p.roundRect(1 * cm, y - 1.4 * cm, 19 * cm, 1.6 * cm, 8, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 13)
    p.drawString(1.5 * cm, y - 0.65 * cm, "NET PAY")
    p.setFont(_FONT_BOLD, 22)
    p.drawRightString(19.5 * cm, y - 0.7 * cm,
                      f"{float(line.net):,.2f} {currency_name_ar(employee.company.base_currency)}")

    # Footer
    p.setFillColor(colors.HexColor("#94A3B8"))
    p.setFont(_FONT_REGULAR, 8)
    foot = f"Generated by Marsoud  ·  {employee.company.name}"
    if employee.company.tax_number:
        foot += f"  ·  VAT # {employee.company.tax_number}"
    p.drawCentredString(10.5 * cm, 1.2 * cm, ar(foot))

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_journal_entry_pdf(entry):
    """Single journal entry PDF — Arabic layout (MARSOUD-JOURNAL-
    EXPORT-AR, Batch 6 Ticket 5, 2026-07-29).

    Columns: الحساب | البيان | مدين | دائن. Amounts stay
    LTR-formatted with thousands separator + 2 decimals so
    accountants can copy them cleanly."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    title = ar(f"قيد يومية — {entry.number or '#' + str(entry.id)}")
    _pdf_header(p, entry.company, title,
                ar(f"التاريخ: {entry.date}  ·  {currency_name_ar(entry.currency)}"))

    y = 24 * cm
    p.setFillColor(colors.HexColor("#475569"))
    p.setFont(_FONT_REGULAR, 10)
    p.drawRightString(19.5 * cm, y,
                      ar(f"الوصف: {entry.description or ''}"))
    y -= 0.5 * cm
    if entry.reference:
        p.drawRightString(19.5 * cm, y,
                          ar(f"المرجع: {entry.reference}"))
        y -= 0.5 * cm
    status_label = "نشط" if entry.is_active else "موقوف"
    p.drawRightString(19.5 * cm, y, ar(f"الحالة: {status_label}"))
    y -= 1 * cm

    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 10)
    # RTL column order (right → left): الحساب | البيان | مدين | دائن.
    p.drawRightString(19.5 * cm, y - 0.2 * cm, ar("الحساب"))
    p.drawRightString(13 * cm, y - 0.2 * cm, ar("البيان"))
    p.drawRightString(7 * cm, y - 0.2 * cm, ar("مدين"))
    p.drawRightString(3.5 * cm, y - 0.2 * cm, ar("دائن"))
    y -= 0.9 * cm

    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 9)
    for line in entry.lines:
        if y < 3 * cm:
            p.showPage()
            y = 27 * cm
        acc_name = line.account.name_ar or line.account.name
        acc_label = f"{line.account.code} · {acc_name[:30]}"
        p.drawRightString(19.5 * cm, y, ar(acc_label))
        p.drawRightString(13 * cm, y, ar((line.memo or "")[:35]))
        p.drawRightString(7 * cm, y, f"{float(line.debit):,.2f}")
        p.drawRightString(3.5 * cm, y, f"{float(line.credit):,.2f}")
        y -= 0.5 * cm

    y -= 0.3 * cm
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.8 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 11)
    p.drawRightString(19.5 * cm, y - 0.15 * cm, ar("الإجمالي"))
    p.drawRightString(7 * cm, y - 0.15 * cm, f"{entry.total_debit:,.2f}")
    p.drawRightString(3.5 * cm, y - 0.15 * cm, f"{entry.total_credit:,.2f}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_journal_entry_excel(entry):
    """Single journal entry Excel — Arabic layout (MARSOUD-JOURNAL-
    EXPORT-AR, Batch 6 Ticket 5, 2026-07-29). RTL sheet view +
    Arabic headers + Arabic currency amount format."""
    wb = Workbook()
    ws = wb.active
    ws.title = (entry.number or f"JE-{entry.id}")[:31]
    # Flip the whole worksheet RTL so column A lands on the right.
    ws.sheet_view.rightToLeft = True
    _excel_styled_header(ws, f"قيد يومية {entry.number or entry.id}",
                         entry.company.name, f"التاريخ: {entry.date}")

    row = 5
    ws.cell(row=row, column=1, value="الوصف:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=entry.description)
    row += 1
    if entry.reference:
        ws.cell(row=row, column=1,
                value="المرجع:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=entry.reference)
        row += 1
    ws.cell(row=row, column=1,
            value="الحالة:").font = Font(bold=True)
    ws.cell(row=row, column=2,
            value="نشط" if entry.is_active else "موقوف")
    row += 2

    for col, h in enumerate(["كود الحساب", "اسم الحساب", "البيان",
                              "مدين", "دائن"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    row += 1
    for line in entry.lines:
        ws.cell(row=row, column=1, value=line.account.code)
        ws.cell(row=row, column=2,
                value=line.account.name_ar or line.account.name)
        ws.cell(row=row, column=3, value=line.memo or "")
        ws.cell(row=row, column=4,
                value=float(line.debit)).number_format = "#,##0.00"
        ws.cell(row=row, column=5,
                value=float(line.credit)).number_format = "#,##0.00"
        row += 1

    ws.cell(row=row, column=1,
            value="الإجمالي").font = Font(bold=True, color="FFFFFF")
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = PatternFill("solid",
                                                          fgColor="2563EB")
    ws.cell(row=row, column=4,
            value=entry.total_debit).number_format = "#,##0.00"
    ws.cell(row=row, column=5,
            value=entry.total_credit).number_format = "#,##0.00"
    ws.cell(row=row, column=4).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=5).font = Font(bold=True, color="FFFFFF")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_journals_list_pdf(company, entries, period_label=""):
    """Filtered list of journals to PDF — Arabic RTL layout
    (MARSOUD-JOURNAL-EXPORT-AR, Batch 6 Ticket 5, 2026-07-29)."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, company, ar("كشف قيود اليومية"),
                 ar(period_label or "كل القيود المفلترة"))

    y = 24.5 * cm
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 9)
    # RTL header order (right → left): الرقم | التاريخ | الوصف |
    # المرجع | الحالة | مدين | دائن.
    p.drawRightString(19.5 * cm, y - 0.2 * cm, ar("الرقم"))
    p.drawRightString(17 * cm, y - 0.2 * cm, ar("التاريخ"))
    p.drawRightString(13 * cm, y - 0.2 * cm, ar("الوصف"))
    p.drawRightString(8.5 * cm, y - 0.2 * cm, ar("المرجع"))
    p.drawRightString(6.5 * cm, y - 0.2 * cm, ar("الحالة"))
    p.drawRightString(4.5 * cm, y - 0.2 * cm, ar("مدين"))
    p.drawRightString(2 * cm, y - 0.2 * cm, ar("دائن"))
    y -= 0.8 * cm

    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 8)
    total_d = total_c = 0.0
    for e in entries:
        if y < 3 * cm:
            p.showPage()
            y = 27 * cm
        p.drawRightString(19.5 * cm, y, ar((e.number or f"#{e.id}")[:10]))
        p.drawRightString(17 * cm, y, str(e.date))
        p.drawRightString(13 * cm, y, ar((e.description or "")[:42]))
        p.drawRightString(8.5 * cm, y, ar((e.reference or "")[:10]))
        p.drawRightString(6.5 * cm, y, ar("نشط" if e.is_active else "موقوف"))
        p.drawRightString(4.5 * cm, y, f"{e.total_debit:,.2f}")
        p.drawRightString(2 * cm, y, f"{e.total_credit:,.2f}")
        total_d += e.total_debit
        total_c += e.total_credit
        y -= 0.45 * cm

    y -= 0.3 * cm
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.8 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 10)
    p.drawRightString(19.5 * cm, y - 0.15 * cm,
                       ar(f"الإجمالي ({len(entries)} قيد)"))
    p.drawRightString(4.5 * cm, y - 0.15 * cm, f"{total_d:,.2f}")
    p.drawRightString(2 * cm, y - 0.15 * cm, f"{total_c:,.2f}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_payroll_run_pdf(run):
    """Full monthly payroll run PDF — one row per employee."""
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, run.company, f"Payroll — {run.period_month}/{run.period_year}",
                f"{run.number or ''}")

    y = 24.5 * cm
    # Column header
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 9)
    p.drawString(1.3 * cm, y - 0.2 * cm, "EMPLOYEE")
    p.drawString(7.5 * cm, y - 0.2 * cm, "DAYS")
    p.drawString(9 * cm, y - 0.2 * cm, "BASIC")
    p.drawString(11.3 * cm, y - 0.2 * cm, "ALLOW")
    p.drawString(13.3 * cm, y - 0.2 * cm, "BONUS")
    p.drawString(15.3 * cm, y - 0.2 * cm, "DEDUCT")
    p.drawRightString(19.5 * cm, y - 0.2 * cm, "NET")
    y -= 0.9 * cm

    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 9)
    for line in run.lines:
        if y < 3 * cm:
            p.showPage()
            y = 27 * cm
        p.drawString(1.3 * cm, y, ar(line.employee.name[:30]))
        p.drawString(7.5 * cm, y, f"{line.working_days}/30")
        p.drawString(9 * cm, y, f"{float(line.basic):,.2f}")
        p.drawString(11.3 * cm, y, f"{float(line.allowances):,.2f}")
        bonus_total = float(line.overtime or 0) + float(line.bonus or 0)
        p.drawString(13.3 * cm, y, f"{bonus_total:,.2f}")
        deduct_total = float(line.deductions or 0) + float(line.absence_deduction or 0) + float(line.late_deduction or 0) + float(line.advance_deduction or 0)
        p.drawString(15.3 * cm, y, f"{deduct_total:,.2f}")
        p.drawRightString(19.5 * cm, y, f"{float(line.net):,.2f}")
        y -= 0.5 * cm

    y -= 0.3 * cm
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.9 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 12)
    p.drawString(1.3 * cm, y - 0.15 * cm, "TOTAL NET")
    p.drawRightString(19.5 * cm, y - 0.15 * cm, f"{float(run.total_net):,.2f} {currency_name_ar(run.company.base_currency)}")

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def export_payroll_run_excel(run):
    """Full monthly payroll run as Excel — auditable detail per employee."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{run.period_year}-{run.period_month:02d}"
    _excel_styled_header(ws, f"Payroll {run.period_month}/{run.period_year}",
                         run.company.name, run.number or "")

    headers = ["Employee #", "Name", "Days", "Basic (prorated)", "Allowances",
               "Overtime", "Bonus", "Fixed Deductions", "Absence", "Late", "Advance", "Net"]
    row = 5
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
        c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    for col_letter in "CDEFGHIJKL":
        ws.column_dimensions[col_letter].width = 14

    row += 1
    for line in run.lines:
        emp = line.employee
        ws.cell(row=row, column=1, value=emp.employee_number or "")
        ws.cell(row=row, column=2, value=emp.name)
        ws.cell(row=row, column=3, value=line.working_days)
        for col_idx, val in enumerate([
            float(line.basic), float(line.allowances), float(line.overtime),
            float(line.bonus), float(line.deductions), float(line.absence_deduction),
            float(line.late_deduction), float(line.advance_deduction), float(line.net),
        ], start=4):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.number_format = "#,##0.00"
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="2563EB")
    for col_idx in range(2, 12):
        ws.cell(row=row, column=col_idx).fill = PatternFill("solid", fgColor="2563EB")
    net_cell = ws.cell(row=row, column=12, value=float(run.total_net))
    net_cell.font = Font(bold=True, color="FFFFFF")
    net_cell.fill = PatternFill("solid", fgColor="2563EB")
    net_cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Generic helpers ────────────────────────────────────────────────────

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _list_pdf(company, title, period_label, headers, rows, totals_row=None, col_widths=None):
    """Generic list-style PDF: title, table, optional totals row.

    headers: list of (label, align) where align is 'left'|'right'
    rows: list of lists matching headers length; values formatted as strings
    totals_row: same shape as a row, displayed bold on a navy background
    col_widths: list of cm widths matching headers length (defaults to even split)
    """
    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, company, title, period_label)

    n = len(headers)
    if not col_widths:
        col_widths = [19.0 / n] * n
    # Compute right-edge x for each column (rtl-ish from left edge)
    x_starts = [1 * cm]
    for w in col_widths[:-1]:
        x_starts.append(x_starts[-1] + w * cm)

    y = 24.5 * cm
    # Column header
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.4 * cm, 19 * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 9)
    for i, (label, align) in enumerate(headers):
        if align == "right":
            p.drawRightString(x_starts[i] + col_widths[i] * cm - 0.2 * cm, y - 0.2 * cm, ar(label))
        else:
            p.drawString(x_starts[i] + 0.2 * cm, y - 0.2 * cm, ar(label))
    y -= 0.9 * cm

    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 9)
    for row in rows:
        if y < 3 * cm:
            p.showPage()
            _pdf_header(p, company, title + " (cont.)", period_label)
            y = 24.5 * cm
        for i, val in enumerate(row):
            align = headers[i][1]
            if align == "right":
                p.drawRightString(x_starts[i] + col_widths[i] * cm - 0.2 * cm, y, ar(val))
            else:
                p.drawString(x_starts[i] + 0.2 * cm, y, ar(val))
        y -= 0.5 * cm

    if totals_row:
        y -= 0.3 * cm
        p.setFillColor(NAVY)
        p.rect(1 * cm, y - 0.5 * cm, 19 * cm, 0.8 * cm, fill=1, stroke=0)
        p.setFillColor(colors.white)
        p.setFont(_FONT_BOLD, 10)
        for i, val in enumerate(totals_row):
            align = headers[i][1]
            if align == "right":
                p.drawRightString(x_starts[i] + col_widths[i] * cm - 0.2 * cm, y - 0.15 * cm, ar(val))
            else:
                p.drawString(x_starts[i] + 0.2 * cm, y - 0.15 * cm, ar(val))

    p.showPage()
    p.save()
    buf.seek(0)
    return buf


def _list_excel(company, title, period_label, headers, rows, totals_row=None):
    """Generic list-style Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    _excel_styled_header(ws, title, company.name, period_label)

    row_n = 5
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row_n, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + col)].width = max(14, len(str(h)) + 4)

    row_n += 1
    for row in rows:
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_n, column=col, value=val)
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
        row_n += 1

    if totals_row:
        for col, val in enumerate(totals_row, 1):
            cell = ws.cell(row=row_n, column=col, value=val)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Income Summary ─────────────────────────────────────────────────────
def export_income_summary(company, fmt, start, end):
    data = income_summary(company.id, start_date=start, end_date=end)
    headers = [("Code", "left"), ("Account", "left"), ("Amount", "right")]
    rows = [[r["code"], r["name"], f"{r['balance']:,.2f}"] for r in data["rows"]]
    totals = ["", "TOTAL REVENUE", f"{data['total']:,.2f}"]
    period = f"{start} → {end}"
    if fmt == "pdf":
        return _list_pdf(company, "Income Summary", period, headers, rows, totals,
                         col_widths=[3, 12, 4]), f"income-summary-{start}-{end}.pdf", "application/pdf"
    return _list_excel(company, "Income Summary", period,
                       ["الكود", "الحساب", "المبلغ"],
                       [[r["code"], r["name"], r["balance"]] for r in data["rows"]],
                       ["", "إجمالي الإيرادات", data["total"]]), \
        f"income-summary-{start}-{end}.xlsx", XLSX_MIME


# ─── Cash Flow ──────────────────────────────────────────────────────────
def export_cash_flow(company, fmt, start, end):
    data = cash_flow(company.id, start_date=start, end_date=end)
    period = f"{start} → {end}" if start else f"→ {end}"
    if fmt == "pdf":
        headers = [("Activity", "left"), ("Net Cash Flow", "right")]
        rows = [
            ["Operating Activities", f"{data['operating']:,.2f}"],
            ["Investing Activities", f"{data['investing']:,.2f}"],
            ["Financing Activities", f"{data['financing']:,.2f}"],
        ]
        totals = ["Net Change in Cash", f"{data['net_change']:,.2f}"]
        return _list_pdf(company, "Cash Flow Statement", period, headers, rows, totals,
                         col_widths=[10, 6]), f"cash-flow-{start}-{end}.pdf", "application/pdf"
    return _list_excel(company, "Cash Flow Statement", period,
                       ["النشاط", "صافي التدفق النقدي"],
                       [["الأنشطة التشغيلية", data["operating"]],
                        ["الأنشطة الاستثمارية", data["investing"]],
                        ["الأنشطة التمويلية", data["financing"]]],
                       ["صافي التغير في النقد", data["net_change"]]), \
        f"cash-flow-{start}-{end}.xlsx", XLSX_MIME


# ─── Expenses Summary ───────────────────────────────────────────────────
def export_expenses_summary(company, fmt, start, end):
    data = expenses_summary(company.id, start_date=start, end_date=end)
    period = f"{start} → {end}"
    if fmt == "pdf":
        headers = [("Code", "left"), ("Account", "left"), ("Entries", "right"), ("Amount", "right")]
        rows = [[r["code"], r["name"], str(r["entry_count"]), f"{r['balance']:,.2f}"] for r in data["rows"]]
        totals = ["", "TOTAL EXPENSES", "", f"{data['total']:,.2f}"]
        return _list_pdf(company, "Expenses Summary", period, headers, rows, totals,
                         col_widths=[3, 10, 2, 4]), f"expenses-summary-{start}-{end}.pdf", "application/pdf"
    return _list_excel(company, "Expenses Summary", period,
                       ["الكود", "الحساب", "عدد القيود", "المبلغ"],
                       [[r["code"], r["name"], r["entry_count"], r["balance"]] for r in data["rows"]],
                       ["", "إجمالي المصروفات", "", data["total"]]), \
        f"expenses-summary-{start}-{end}.xlsx", XLSX_MIME


# ─── P&L Compared ───────────────────────────────────────────────────────
def export_pl_compared(company, fmt, start, end):
    data = income_statement_compared(company.id, start_date=start, end_date=end)
    cur_label = f"{start.isoformat()} → {end.isoformat()}"
    prior_label = f"{data['prior_start'].isoformat()} → {data['prior_end'].isoformat()}"
    period = f"Compared: {cur_label} vs {prior_label}"
    rows_data = [
        ("Revenue", data["current"]["total_revenue"], data["prior"]["total_revenue"], data["delta_revenue"]),
        ("Expenses", data["current"]["total_expense"], data["prior"]["total_expense"], data["delta_expense"]),
        ("Net Profit", data["current"]["net_income"], data["prior"]["net_income"], data["delta_net"]),
    ]
    if fmt == "pdf":
        headers = [("Line", "left"), (cur_label, "right"), (prior_label, "right"), ("Δ", "right")]
        rows = [[label, f"{cur:,.2f}", f"{pri:,.2f}", f"{delta:+,.2f}"] for label, cur, pri, delta in rows_data]
        return _list_pdf(company, "P&L Compared (vs Prior Year)", period, headers, rows,
                         col_widths=[5, 5, 5, 4]), f"pl-compared-{start}-{end}.pdf", "application/pdf"
    return _list_excel(company, "P&L Compared", period,
                       ["البند", "الفترة الحالية", "الفترة السابقة", "التغير"],
                       [[label, cur, pri, delta] for label, cur, pri, delta in rows_data]), \
        f"pl-compared-{start}-{end}.xlsx", XLSX_MIME


# ─── AR Aging ───────────────────────────────────────────────────────────
def export_ar_aging(company, fmt, end):
    data = aging_report(company.id, as_of=end)
    rows_data = [(r["customer"], r["current"], r["d30"], r["d60"], r["d90"], r["d90plus"], r["total"])
                 for r in data["rows"]]
    t = data["totals"]
    period = f"As of {end}"
    if fmt == "pdf":
        headers = [("Customer", "left"), ("Current", "right"), ("1-30", "right"),
                   ("31-60", "right"), ("61-90", "right"), ("90+", "right"), ("Total", "right")]
        rows = [[name, f"{c:,.2f}", f"{d30:,.2f}", f"{d60:,.2f}", f"{d90:,.2f}", f"{d90p:,.2f}", f"{tot:,.2f}"]
                for name, c, d30, d60, d90, d90p, tot in rows_data]
        totals = ["TOTAL", f"{t['current']:,.2f}", f"{t['d30']:,.2f}", f"{t['d60']:,.2f}",
                  f"{t['d90']:,.2f}", f"{t['d90plus']:,.2f}", f"{t['total']:,.2f}"]
        return _list_pdf(company, "Accounts Receivable Aging", period, headers, rows, totals,
                         col_widths=[5, 2.5, 2.5, 2.5, 2.5, 2.5, 1.5]), \
            f"ar-aging-{end}.pdf", "application/pdf"
    return _list_excel(company, "AR Aging", period,
                       ["العميل", "جاري", "1-30", "31-60", "61-90", "90+", "الإجمالي"],
                       [list(r) for r in rows_data],
                       ["الإجمالي", t["current"], t["d30"], t["d60"], t["d90"], t["d90plus"], t["total"]]), \
        f"ar-aging-{end}.xlsx", XLSX_MIME


# ─── AP Aging ───────────────────────────────────────────────────────────
def export_ap_aging(company, fmt, end):
    data = ap_aging_report(company.id, as_of=end)
    rows_data = [(r["vendor"], r["current"], r["d30"], r["d60"], r["d90"], r["d90plus"], r["total"])
                 for r in data["rows"]]
    t = data["totals"]
    period = f"As of {end}"
    if fmt == "pdf":
        headers = [("Vendor", "left"), ("Current", "right"), ("1-30", "right"),
                   ("31-60", "right"), ("61-90", "right"), ("90+", "right"), ("Total", "right")]
        rows = [[name, f"{c:,.2f}", f"{d30:,.2f}", f"{d60:,.2f}", f"{d90:,.2f}", f"{d90p:,.2f}", f"{tot:,.2f}"]
                for name, c, d30, d60, d90, d90p, tot in rows_data]
        totals = ["TOTAL", f"{t['current']:,.2f}", f"{t['d30']:,.2f}", f"{t['d60']:,.2f}",
                  f"{t['d90']:,.2f}", f"{t['d90plus']:,.2f}", f"{t['total']:,.2f}"]
        return _list_pdf(company, "Accounts Payable Aging", period, headers, rows, totals,
                         col_widths=[5, 2.5, 2.5, 2.5, 2.5, 2.5, 1.5]), \
            f"ap-aging-{end}.pdf", "application/pdf"
    return _list_excel(company, "AP Aging", period,
                       ["المورد", "جاري", "1-30", "31-60", "61-90", "90+", "الإجمالي"],
                       [list(r) for r in rows_data],
                       ["الإجمالي", t["current"], t["d30"], t["d60"], t["d90"], t["d90plus"], t["total"]]), \
        f"ap-aging-{end}.xlsx", XLSX_MIME


# ─── VAT Report ─────────────────────────────────────────────────────────
def export_vat_report(company, fmt, start, end):
    data = vat_report(company.id, start_date=start, end_date=end)
    period = f"{start} → {end}"
    if fmt == "pdf":
        # Special gov-ready layout — single-page summary
        buf = io.BytesIO()
        p = canvas.Canvas(buf, pagesize=A4)
        _pdf_header(p, company, "VAT Return", period)
        y = 24 * cm
        p.setFillColor(colors.HexColor("#475569"))
        p.setFont(_FONT_REGULAR, 11)
        if company.tax_number:
            p.drawString(1.5 * cm, y, ar(f"Tax Number: {company.tax_number}"))
            y -= 0.8 * cm

        # Three rows: collected, paid, net — large summary cards
        for label, amount, color in [
            ("VAT Collected (from sales)", data["collected"], colors.HexColor("#10B981")),
            ("VAT Paid (to suppliers)", data["paid"], colors.HexColor("#F59E0B")),
        ]:
            p.setFillColor(color)
            p.rect(1 * cm, y - 1.5 * cm, 19 * cm, 1.2 * cm, fill=1, stroke=0)
            p.setFillColor(colors.white)
            p.setFont(_FONT_BOLD, 12)
            p.drawString(1.5 * cm, y - 0.8 * cm, ar(label))
            p.setFont(_FONT_BOLD, 16)
            p.drawRightString(19.5 * cm, y - 0.8 * cm, f"{amount:,.2f} {currency_name_ar(company.base_currency)}")
            y -= 2 * cm

        p.setFillColor(NAVY)
        p.rect(1 * cm, y - 1.5 * cm, 19 * cm, 1.4 * cm, fill=1, stroke=0)
        p.setFillColor(colors.white)
        p.setFont(_FONT_BOLD, 13)
        p.drawString(1.5 * cm, y - 0.8 * cm, ar("NET DUE TO GOVERNMENT"))
        p.setFont(_FONT_BOLD, 18)
        p.drawRightString(19.5 * cm, y - 0.8 * cm, f"{data['net']:,.2f} {currency_name_ar(company.base_currency)}")

        p.showPage()
        p.save()
        buf.seek(0)
        return buf, f"vat-{start}-{end}.pdf", "application/pdf"

    return _list_excel(company, "VAT Report", period,
                       ["البند", "المبلغ"],
                       [["الضريبة المحصلة", data["collected"]],
                        ["الضريبة المدفوعة", data["paid"]],
                        ["الصافي المستحق", data["net"]]]), \
        f"vat-{start}-{end}.xlsx", XLSX_MIME


# ─── Payroll Summary ────────────────────────────────────────────────────
def export_payroll_summary(company, fmt, year=None, month=None):
    data = payroll_summary_report(company.id, year=year, month=month)
    rows_data = [(r["period"], r["run_number"], r["employee"], r["basic"], r["allowances"],
                  r["overtime"], r["bonus"], r["deductions"], r["net"]) for r in data["rows"]]
    t = data["totals"]
    period = f"Year {year or 'all'} · Month {month or 'all'}"
    if fmt == "pdf":
        headers = [("Period", "left"), ("Run", "left"), ("Employee", "left"),
                   ("Basic", "right"), ("Allow", "right"), ("OT", "right"),
                   ("Bonus", "right"), ("Deduct", "right"), ("Net", "right")]
        rows = [[per, run, emp, f"{b:,.2f}", f"{a:,.2f}", f"{ot:,.2f}", f"{bn:,.2f}", f"{d:,.2f}", f"{n:,.2f}"]
                for per, run, emp, b, a, ot, bn, d, n in rows_data]
        totals = ["", "", "TOTAL", f"{t['basic']:,.2f}", f"{t['allowances']:,.2f}",
                  f"{t['overtime']:,.2f}", f"{t['bonus']:,.2f}", f"{t['deductions']:,.2f}", f"{t['net']:,.2f}"]
        return _list_pdf(company, "Payroll Summary", period, headers, rows, totals,
                         col_widths=[1.5, 2, 3.5, 2, 2, 1.5, 1.5, 2, 2]), \
            f"payroll-summary.pdf", "application/pdf"
    return _list_excel(company, "Payroll Summary", period,
                       ["الفترة", "الكشف", "الموظف", "الأساسي", "البدلات", "أوفرتايم", "بونص", "خصومات", "الصافي"],
                       [list(r) for r in rows_data],
                       ["", "", "الإجمالي", t["basic"], t["allowances"], t["overtime"], t["bonus"], t["deductions"], t["net"]]), \
        f"payroll-summary.xlsx", XLSX_MIME


# ─── Fixed Assets Report ────────────────────────────────────────────────
def export_fixed_assets(company, fmt):
    data = fixed_assets_report(company.id)
    rows_data = [(r["name"], r["vendor"] or "—", str(r["purchase_date"]),
                  f"{r['useful_life_years']}y", r["cost"], r["annual_dep"],
                  r["accumulated_dep"], r["nbv"]) for r in data["rows"]]
    t = data["totals"]
    period = "All active fixed assets"
    if fmt == "pdf":
        headers = [("Asset", "left"), ("Vendor", "left"), ("Purchase", "left"),
                   ("Life", "left"), ("Cost", "right"), ("Annual Dep", "right"),
                   ("Acc. Dep", "right"), ("NBV", "right")]
        rows = [[n, v, p, l, f"{c:,.2f}", f"{ad:,.2f}", f"{accd:,.2f}", f"{nbv:,.2f}"]
                for n, v, p, l, c, ad, accd, nbv in rows_data]
        totals = ["TOTAL", "", "", "", f"{t['cost']:,.2f}", f"{t['annual_dep']:,.2f}",
                  f"{t['accumulated_dep']:,.2f}", f"{t['nbv']:,.2f}"]
        return _list_pdf(company, "Fixed Assets Report", period, headers, rows, totals,
                         col_widths=[3, 2.5, 2, 1.2, 2.5, 2.5, 2.5, 2.8]), \
            f"fixed-assets.pdf", "application/pdf"
    return _list_excel(company, "Fixed Assets", period,
                       ["الأصل", "المورد", "تاريخ الشراء", "العمر", "التكلفة", "إهلاك سنوي", "مجمع الإهلاك", "القيمة الدفترية"],
                       [list(r) for r in rows_data],
                       ["الإجمالي", "", "", "", t["cost"], t["annual_dep"], t["accumulated_dep"], t["nbv"]]), \
        f"fixed-assets.xlsx", XLSX_MIME


# ─── Dispatcher ─────────────────────────────────────────────────────────
def export_report(company, report_type, fmt, start, end, **kwargs):
    """Dispatch to the right export function. fmt = 'pdf' or 'excel'."""
    if report_type == "balance-sheet":
        if fmt == "pdf":
            return export_balance_sheet_pdf(company, end), f"balance-sheet-{end}.pdf", "application/pdf"
        return export_balance_sheet_excel(company, end), f"balance-sheet-{end}.xlsx", XLSX_MIME
    if report_type == "income-statement":
        if fmt == "pdf":
            return export_income_statement_pdf(company, start, end), f"income-statement-{start}-{end}.pdf", "application/pdf"
        return export_income_statement_excel(company, start, end), f"income-statement-{start}-{end}.xlsx", XLSX_MIME
    if report_type == "cash-flow":
        return export_cash_flow(company, fmt, start, end)
    if report_type == "income-summary":
        return export_income_summary(company, fmt, start, end)
    if report_type == "expenses-summary":
        return export_expenses_summary(company, fmt, start, end)
    if report_type == "pl-compared":
        return export_pl_compared(company, fmt, start, end)
    if report_type == "ar-aging":
        return export_ar_aging(company, fmt, end)
    if report_type == "ap-aging":
        return export_ap_aging(company, fmt, end)
    if report_type == "vat":
        return export_vat_report(company, fmt, start, end)
    if report_type == "payroll-summary":
        return export_payroll_summary(company, fmt,
                                      year=kwargs.get("year"), month=kwargs.get("month"))
    if report_type == "fixed-assets":
        return export_fixed_assets(company, fmt)
    if report_type == "low-stock":
        if fmt == "pdf":
            return (export_low_stock_pdf(company),
                    f"low-stock-{date.today()}.pdf", "application/pdf")
        return (export_low_stock_excel(company),
                f"low-stock-{date.today()}.xlsx", XLSX_MIME)
    if report_type == "stock-movements":
        if fmt == "pdf":
            return (export_stock_movements_pdf(company, start, end),
                    f"stock-movements-{start}-{end}.pdf", "application/pdf")
        return (export_stock_movements_excel(company, start, end),
                f"stock-movements-{start}-{end}.xlsx", XLSX_MIME)
    if report_type == "inventory-balance":
        wh_id = kwargs.get("warehouse_id")
        if fmt == "pdf":
            return (export_inventory_balance_pdf(company, warehouse_id=wh_id),
                    f"inventory-balance-{date.today()}.pdf", "application/pdf")
        return (export_inventory_balance_excel(company, warehouse_id=wh_id),
                f"inventory-balance-{date.today()}.xlsx", XLSX_MIME)
    if report_type == "profitability":
        if fmt == "pdf":
            return (export_profitability_pdf(company, start, end),
                    f"profitability-{start}-{end}.pdf", "application/pdf")
        return (export_profitability_excel(company, start, end),
                f"profitability-{start}-{end}.xlsx", XLSX_MIME)
    if report_type == "cashier-sales":
        if fmt == "pdf":
            return (export_cashier_sales_pdf(company, start, end),
                    f"cashier-sales-{start}-{end}.pdf", "application/pdf")
        return (export_cashier_sales_excel(company, start, end),
                f"cashier-sales-{start}-{end}.xlsx", XLSX_MIME)
    raise ValueError(f"Unknown report: {report_type}")


# ─── ERP-02 — inventory exports ──────────────────────────────────────────
def export_low_stock_excel(company):
    """Variants whose total qty is below their reorder_level."""
    from app.services.inventory import low_stock_variants
    wb = Workbook()
    ws = wb.active
    ws.title = "Low Stock"
    _excel_styled_header(ws, "أصناف تحت حد الطلب", company.name,
                         f"كما في {date.today()}")
    row = 5
    headers = ["SKU", "المنتج", "المتاح", "حد الطلب", "متوسط التكلفة", "القيمة الدفترية"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for v in low_stock_variants(company.id):
        ws.cell(row=row, column=1, value=v.sku)
        ws.cell(row=row, column=2, value=v.display_name)
        ws.cell(row=row, column=3, value=v.total_qty).number_format = "#,##0.00"
        ws.cell(row=row, column=4, value=float(v.reorder_level or 0)).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=v.average_cost).number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=v.total_value).number_format = "#,##0.00"
        row += 1
    for i, w in enumerate([14, 35, 12, 12, 14, 16], start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_stock_movements_excel(company, start, end):
    """Full movement log for the given date range."""
    from app.models import StockMovement, StockMovementKind
    from datetime import datetime as _dt, timedelta as _td
    rows = StockMovement.query.filter(
        StockMovement.company_id == company.id,
        StockMovement.created_at >= _dt.combine(start, _dt.min.time()),
        StockMovement.created_at < _dt.combine(end, _dt.min.time()) + _td(days=1),
    ).order_by(StockMovement.created_at.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Movements"
    _excel_styled_header(ws, "سجل حركات المخزون", company.name,
                         f"من {start} إلى {end}")
    row = 5
    headers = ["التاريخ", "النوع", "الصنف", "المخزن", "الكمية",
               "تكلفة الوحدة", "الرصيد بعد", "المصدر", "المنفّذ", "السبب"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    kind_labels = {k.value: k.label_ar for k in StockMovementKind}
    from app.services.time import to_company_tz_str
    for m in rows:
        ws.cell(row=row, column=1,
                 value=to_company_tz_str(m.created_at, company,
                                          "%Y-%m-%d %H:%M") or "")
        ws.cell(row=row, column=2, value=kind_labels.get(m.kind, m.kind))
        ws.cell(row=row, column=3, value=m.variant.sku if m.variant else "")
        ws.cell(row=row, column=4, value=m.warehouse.code if m.warehouse else "")
        ws.cell(row=row, column=5, value=float(m.qty_delta or 0)).number_format = "+#,##0.00;-#,##0.00"
        ws.cell(row=row, column=6, value=float(m.unit_cost_at_time or 0)).number_format = "#,##0.0000"
        ws.cell(row=row, column=7, value=float(m.balance_qty_after or 0)).number_format = "#,##0.00"
        ws.cell(row=row, column=8,
                value=f"{m.source_type or ''}#{m.source_id}" if m.source_id else (m.source_type or ""))
        ws.cell(row=row, column=9, value=m.actor.full_name if m.actor else "نظام")
        ws.cell(row=row, column=10, value=m.reason or "")
        row += 1
    for i, w in enumerate([18, 14, 14, 10, 12, 12, 12, 16, 18, 30], start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── GAP closure: rest of the ERP-01 reports (PDF + Excel) ──────────────
def _simple_pdf_table(buf, company, title, period, headers, rows,
                      col_widths=None):
    """Render an A4 portrait PDF with the Marsoud Arabic header + a single
    full-width table. Auto-paginates when the page fills. Numeric columns
    are right-aligned; text columns are left-aligned in their cells.

    headers: list[str].
    rows: list[list[str]]. Caller pre-formats numbers as strings.
    col_widths: list[float] in cm; defaults to equal columns.
    """
    p = canvas.Canvas(buf, pagesize=A4)
    _pdf_header(p, company, title, period)
    y = 24.5 * cm
    n = len(headers)
    if col_widths is None:
        total = 19.0
        col_widths = [total / n] * n
    # cumulative x offsets in cm (table starts at 1cm)
    x_starts = [1.0]
    for w in col_widths:
        x_starts.append(x_starts[-1] + w)
    # Header band
    p.setFillColor(NAVY)
    p.rect(1 * cm, y - 0.5 * cm, sum(col_widths) * cm, 0.7 * cm, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont(_FONT_BOLD, 9)
    for i, h in enumerate(headers):
        cx = (x_starts[i] + col_widths[i] / 2) * cm
        p.drawCentredString(cx, y - 0.25 * cm, ar(h))
    y -= 1 * cm
    p.setFillColor(colors.black)
    p.setFont(_FONT_REGULAR, 9)
    for row in rows:
        if y < 2.5 * cm:
            p.showPage()
            _pdf_header(p, company, title + " (cont.)", period)
            y = 24.5 * cm
            # Re-draw header band
            p.setFillColor(NAVY)
            p.rect(1 * cm, y - 0.5 * cm, sum(col_widths) * cm, 0.7 * cm, fill=1, stroke=0)
            p.setFillColor(colors.white)
            p.setFont(_FONT_BOLD, 9)
            for i, h in enumerate(headers):
                cx = (x_starts[i] + col_widths[i] / 2) * cm
                p.drawCentredString(cx, y - 0.25 * cm, ar(h))
            y -= 1 * cm
            p.setFillColor(colors.black)
            p.setFont(_FONT_REGULAR, 9)
        for i, cell in enumerate(row):
            cell_s = str(cell)
            # Right-align if it looks numeric (digits, comma, dot, %).
            looks_num = bool(cell_s) and all(
                c.isdigit() or c in ".,%-+" for c in cell_s
            )
            cx_right = (x_starts[i] + col_widths[i] - 0.1) * cm
            cx_left = (x_starts[i] + 0.1) * cm
            if looks_num:
                p.drawRightString(cx_right, y, cell_s)
            else:
                p.drawString(cx_left, y, ar(cell_s))
        y -= 0.5 * cm
    p.showPage()
    p.save()
    buf.seek(0)
    return buf


# ─── 1. رصيد المخزون الحالي (inventory balance) ────────────────────────
def _inventory_balance_rows(company_id, warehouse_id=None):
    """One row per (variant, warehouse) with current qty + value.

    warehouse_id: optional filter to scope the report to a single
    warehouse (كشف حساب لمخزن واحد).
    """
    from app.models import StockBalance, ProductVariant, Warehouse
    out = []
    q = (
        db.session.query(StockBalance, ProductVariant, Warehouse)
        .join(ProductVariant, StockBalance.variant_id == ProductVariant.id)
        .join(Warehouse, StockBalance.warehouse_id == Warehouse.id)
        .filter(ProductVariant.company_id == company_id)
    )
    if warehouse_id:
        q = q.filter(StockBalance.warehouse_id == warehouse_id)
    rows = q.order_by(ProductVariant.sku, Warehouse.code).all()
    for bal, v, w in rows:
        qty = float(bal.qty or 0)
        value = float(bal.value or 0)
        avg = (value / qty) if qty > 0 else 0
        out.append({
            "sku": v.sku, "name": v.display_name,
            "warehouse": w.code, "qty": qty,
            "avg_cost": avg, "value": value,
        })
    return out


def export_inventory_balance_excel(company, warehouse_id=None):
    rows = _inventory_balance_rows(company.id, warehouse_id=warehouse_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Balance"
    subtitle = f"كما في {date.today()}"
    if warehouse_id and rows:
        subtitle = f"مخزن: {rows[0]['warehouse']} — {subtitle}"
    _excel_styled_header(ws, "رصيد المخزون الحالي", company.name, subtitle)
    row = 5
    headers = ["SKU", "المنتج", "المخزن", "الكمية", "متوسط التكلفة", "القيمة"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    total_value = 0.0
    for r in rows:
        ws.cell(row=row, column=1, value=r["sku"])
        ws.cell(row=row, column=2, value=r["name"])
        ws.cell(row=row, column=3, value=r["warehouse"])
        ws.cell(row=row, column=4, value=r["qty"]).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=r["avg_cost"]).number_format = "#,##0.0000"
        ws.cell(row=row, column=6, value=r["value"]).number_format = "#,##0.00"
        total_value += r["value"]
        row += 1
    ws.cell(row=row, column=5, value="إجمالي القيمة").font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_value).number_format = "#,##0.00"
    ws.cell(row=row, column=6).font = Font(bold=True)
    for i, w in enumerate([14, 30, 12, 12, 14, 14], start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_inventory_balance_pdf(company, warehouse_id=None):
    rows = _inventory_balance_rows(company.id, warehouse_id=warehouse_id)
    out = [
        [r["sku"], r["name"], r["warehouse"],
         f"{r['qty']:,.2f}", f"{r['avg_cost']:,.4f}", f"{r['value']:,.2f}"]
        for r in rows
    ]
    total = sum(r["value"] for r in rows)
    out.append(["", "", "", "", "إجمالي", f"{total:,.2f}"])
    buf = io.BytesIO()
    subtitle = f"كما في {date.today()}"
    if warehouse_id and rows:
        subtitle = f"مخزن: {rows[0]['warehouse']} — {subtitle}"
    return _simple_pdf_table(
        buf, company, "رصيد المخزون الحالي", subtitle,
        ["SKU", "المنتج", "المخزن", "الكمية", "متوسط التكلفة", "القيمة"],
        out, col_widths=[2.5, 6.0, 2.0, 2.5, 3.0, 3.0],
    )


# ─── 2. ربحية المنتج (profitability) ───────────────────────────────────
def _profitability_rows(company_id, start, end):
    from app.models import (
        InvoiceItem, Invoice, InvoiceStatus, ProductVariant,
    )
    items = (
        db.session.query(InvoiceItem, Invoice)
        .join(Invoice, InvoiceItem.invoice_id == Invoice.id)
        .filter(Invoice.company_id == company_id,
                Invoice.issue_date >= start, Invoice.issue_date <= end,
                Invoice.status != InvoiceStatus.DRAFT,
                Invoice.status != InvoiceStatus.VOIDED,
                InvoiceItem.variant_id.isnot(None))
        .all()
    )
    agg = {}
    for it, inv in items:
        if not it.variant:
            continue
        a = agg.setdefault(it.variant_id, {
            "sku": it.variant.sku, "name": it.variant.display_name,
            "qty": 0, "revenue": 0, "cogs": 0,
        })
        a["qty"] += float(it.quantity or 0)
        a["revenue"] += float(it.line_total or 0)
        a["cogs"] += float(it.quantity or 0) * float(it.unit_cost_at_sale or 0)
    out = []
    for r in agg.values():
        gp = r["revenue"] - r["cogs"]
        gm = (gp / r["revenue"] * 100) if r["revenue"] > 0 else 0
        out.append({**r, "gross_profit": gp, "gross_margin": gm})
    out.sort(key=lambda r: -r["gross_profit"])
    return out


def export_profitability_excel(company, start, end):
    rows = _profitability_rows(company.id, start, end)
    wb = Workbook()
    ws = wb.active
    ws.title = "Profitability"
    _excel_styled_header(ws, "ربحية المنتجات", company.name,
                         f"من {start} إلى {end}")
    row = 5
    headers = ["SKU", "المنتج", "الكمية المباعة",
               "المبيعات", "تكلفة البضاعة", "مجمل الربح", "الهامش %"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    total_rev = total_cogs = 0.0
    for r in rows:
        ws.cell(row=row, column=1, value=r["sku"])
        ws.cell(row=row, column=2, value=r["name"])
        ws.cell(row=row, column=3, value=r["qty"]).number_format = "#,##0.00"
        ws.cell(row=row, column=4, value=r["revenue"]).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=r["cogs"]).number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=r["gross_profit"]).number_format = "#,##0.00"
        ws.cell(row=row, column=7, value=r["gross_margin"]).number_format = "0.0\"%\""
        total_rev += r["revenue"]; total_cogs += r["cogs"]
        row += 1
    total_gp = total_rev - total_cogs
    total_gm = (total_gp / total_rev * 100) if total_rev > 0 else 0
    ws.cell(row=row, column=3, value="الإجمالي").font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_rev).number_format = "#,##0.00"
    ws.cell(row=row, column=5, value=total_cogs).number_format = "#,##0.00"
    ws.cell(row=row, column=6, value=total_gp).number_format = "#,##0.00"
    ws.cell(row=row, column=7, value=total_gm).number_format = "0.0\"%\""
    for col in (4, 5, 6, 7):
        ws.cell(row=row, column=col).font = Font(bold=True)
    for i, w in enumerate([12, 28, 12, 14, 14, 14, 10], start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_profitability_pdf(company, start, end):
    rows = _profitability_rows(company.id, start, end)
    out = [
        [r["sku"], r["name"], f"{r['qty']:,.2f}",
         f"{r['revenue']:,.2f}", f"{r['cogs']:,.2f}",
         f"{r['gross_profit']:,.2f}", f"{r['gross_margin']:.1f}%"]
        for r in rows
    ]
    total_rev = sum(r["revenue"] for r in rows)
    total_cogs = sum(r["cogs"] for r in rows)
    total_gp = total_rev - total_cogs
    total_gm = (total_gp / total_rev * 100) if total_rev > 0 else 0
    out.append(["", "إجمالي", "",
                f"{total_rev:,.2f}", f"{total_cogs:,.2f}",
                f"{total_gp:,.2f}", f"{total_gm:.1f}%"])
    buf = io.BytesIO()
    return _simple_pdf_table(
        buf, company, "ربحية المنتجات", f"من {start} إلى {end}",
        ["SKU", "المنتج", "الكمية", "المبيعات",
         "تكلفة البضاعة", "مجمل الربح", "الهامش"],
        out, col_widths=[2.0, 5.0, 2.0, 2.7, 2.7, 2.7, 1.9],
    )


# ─── 3. مبيعات الكاشير (cashier sales) ─────────────────────────────────
def _cashier_sales_rows(company_id, start, end):
    from app.models import Invoice, InvoiceStatus
    invs = Invoice.query.filter(
        Invoice.company_id == company_id,
        Invoice.source == "POS",
        Invoice.issue_date >= start,
        Invoice.issue_date <= end,
    ).all()
    agg = {}
    for inv in invs:
        cid_key = inv.cashier_id or 0
        a = agg.setdefault(cid_key, {
            "cashier": inv.cashier.full_name if inv.cashier else "—",
            "orders": 0, "voids": 0, "gross": 0, "net": 0,
            "by_method": {},
        })
        a["orders"] += 1
        if inv.is_voided:
            a["voids"] += 1
            continue
        a["gross"] += float(inv.total or 0)
        a["net"] += float(inv.total or 0)
        for pay in inv.payments:
            mname = pay.payment_method.name_ar if pay.payment_method else (pay.method or "غير محدد")
            a["by_method"][mname] = a["by_method"].get(mname, 0) + float(pay.amount or 0)
    return sorted(agg.values(), key=lambda r: -r["gross"])


def export_cashier_sales_excel(company, start, end):
    rows = _cashier_sales_rows(company.id, start, end)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cashier Sales"
    _excel_styled_header(ws, "مبيعات الكاشير", company.name,
                         f"من {start} إلى {end}")
    row = 5
    headers = ["الكاشير", "الأوردرات", "الملغى",
               "المبيعات", "صافي", "حسب طريقة الدفع"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
    row += 1
    for r in rows:
        ws.cell(row=row, column=1, value=r["cashier"])
        ws.cell(row=row, column=2, value=r["orders"])
        ws.cell(row=row, column=3, value=r["voids"])
        ws.cell(row=row, column=4, value=r["gross"]).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=r["net"]).number_format = "#,##0.00"
        methods = ", ".join(f"{m}: {v:,.2f}" for m, v in r["by_method"].items())
        ws.cell(row=row, column=6, value=methods)
        row += 1
    for i, w in enumerate([22, 10, 10, 14, 14, 40], start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_cashier_sales_pdf(company, start, end):
    rows = _cashier_sales_rows(company.id, start, end)
    out = [
        [r["cashier"], str(r["orders"]), str(r["voids"]),
         f"{r['gross']:,.2f}", f"{r['net']:,.2f}",
         ", ".join(f"{m}:{v:,.0f}" for m, v in r["by_method"].items())]
        for r in rows
    ]
    buf = io.BytesIO()
    return _simple_pdf_table(
        buf, company, "مبيعات الكاشير", f"من {start} إلى {end}",
        ["الكاشير", "الأوردرات", "الملغى", "المبيعات", "صافي", "طرق الدفع"],
        out, col_widths=[3.5, 1.8, 1.8, 2.8, 2.8, 6.3],
    )


# ─── 4. Low-stock PDF (Excel already exists) ───────────────────────────
def export_low_stock_pdf(company):
    from app.services.inventory import low_stock_variants
    rows = []
    for v in low_stock_variants(company.id):
        rows.append([v.sku, v.display_name,
                     f"{v.total_qty:,.2f}",
                     f"{float(v.reorder_level or 0):,.2f}",
                     f"{v.average_cost:,.2f}",
                     f"{v.total_value:,.2f}"])
    buf = io.BytesIO()
    return _simple_pdf_table(
        buf, company, "أصناف تحت حد الطلب", f"كما في {date.today()}",
        ["SKU", "المنتج", "المتاح", "حد الطلب",
         "متوسط التكلفة", "القيمة"],
        rows, col_widths=[2.5, 6.0, 2.5, 2.5, 2.5, 3.0],
    )


# ─── 5. Movement log PDF (Excel already exists) ────────────────────────
def export_stock_movements_pdf(company, start, end):
    from app.models import StockMovement, StockMovementKind
    from datetime import datetime as _dt, timedelta as _td
    rows = StockMovement.query.filter(
        StockMovement.company_id == company.id,
        StockMovement.created_at >= _dt.combine(start, _dt.min.time()),
        StockMovement.created_at < _dt.combine(end, _dt.min.time()) + _td(days=1),
    ).order_by(StockMovement.created_at.asc()).all()
    kind_labels = {k.value: k.label_ar for k in StockMovementKind}
    out = []
    from app.services.time import to_company_tz_str
    for m in rows:
        out.append([
            to_company_tz_str(m.created_at, company,
                                 "%Y-%m-%d %H:%M") or "",
            kind_labels.get(m.kind, m.kind),
            m.variant.sku if m.variant else "",
            m.warehouse.code if m.warehouse else "",
            f"{float(m.qty_delta or 0):+,.2f}",
            f"{float(m.balance_qty_after or 0):,.2f}",
            m.actor.full_name if m.actor else "نظام",
        ])
    buf = io.BytesIO()
    return _simple_pdf_table(
        buf, company, "سجل حركات المخزون", f"من {start} إلى {end}",
        ["التاريخ", "النوع", "الصنف", "المخزن",
         "الكمية", "الرصيد بعد", "المنفّذ"],
        out, col_widths=[2.8, 2.0, 2.5, 1.5, 2.0, 2.5, 5.7],
    )


# ─── MARSOUD: Excel export of all Leads ─────────────────────────────────
def export_leads_excel(company, leads):
    """Export every field of every lead to a single Excel sheet.

    Abdelhamid's ask: 'I want to export Leads to Excel containing every
    field that's filled in'. Mirrors the lead detail page columns + the
    optional rich-text fields (request_description, sales_action_required,
    notes, meeting_notes, expected_value).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    _excel_styled_header(
        ws, "تقرير العملاء المحتملين",
        company.name, f"عدد السجلات: {len(leads)}",
    )

    headers = [
        "#", "الاسم", "الهاتف", "البريد", "الخدمة المطلوبة",
        "النوع", "المصدر", "الحالة", "المسؤول", "أنشأها",
        "القيمة المتوقعة", "تاريخ الإنشاء", "اجتماع قادم",
        "ملاحظات اجتماع", "وصف الطلب", "المطلوب من السيلز",
        "ملاحظات عامة", "حالة التحويل", "سبب الخسارة",
    ]
    row = 5
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Column widths (tuned for Arabic content)
    widths = [4, 22, 18, 26, 26, 12, 14, 14, 18, 18,
              14, 18, 18, 30, 30, 30, 28, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    row += 1
    for i, L in enumerate(leads, 1):
        rep_name = L.assigned_to.full_name if L.assigned_to else ""
        creator_name = L.created_by.full_name if L.created_by else ""
        status_label = L.status.label_ar if L.status else ""
        type_label = L.lead_type or ""
        source_label = L.source or ""
        expected = float(L.expected_value) if L.expected_value else 0.0
        from app.services.time import to_company_tz_str
        meeting_str = (to_company_tz_str(L.next_meeting, L.company,
                                             "%Y-%m-%d %H:%M") or ""
                       if L.next_meeting else "")
        created_str = (to_company_tz_str(L.created_at, L.company,
                                             "%Y-%m-%d %H:%M") or ""
                       if L.created_at else "")
        converted = "نعم" if L.is_converted else ""

        values = [
            i, L.client_name or "", L.phone or "", L.email or "",
            L.service_needed or "", type_label, source_label,
            status_label, rep_name, creator_name,
            expected, created_str, meeting_str,
            L.meeting_notes or "", L.request_description or "",
            L.sales_action_required or "", L.notes or "",
            converted, L.lost_reason or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 11:   # expected value column
                cell.number_format = "#,##0.00"
        row += 1

    # Totals row for expected_value
    if leads:
        ws.cell(row=row, column=10, value="إجمالي القيمة المتوقعة:").font = Font(bold=True)
        total_expected = sum(float(L.expected_value or 0) for L in leads)
        c = ws.cell(row=row, column=11, value=total_expected)
        c.number_format = "#,##0.00"
        c.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── MARSOUD-LEAD-EXPORT-BY-CAMPAIGN (Abdelhamid 2026-07-15) ────────────
# Alternative shape of the leads export: one sheet per campaign,
# sorted by status inside each sheet, plus a summary sheet up front
# showing count-per-status per campaign. Handles the ticket's ask
# "export by campaign type with categorization by status inside."
def export_leads_by_campaign_excel(company, leads):
    """Same fields as export_leads_excel, but rearranged as:

      · Sheet 1 "ملخص الحملات" — table of {campaign × status → count}
        + a grand total column per campaign + column totals.
      · Sheet N (per campaign) — leads on that campaign, sorted by
        status then created_at. Un-campaigned leads land in a
        "بدون حملة" sheet at the end.
    """
    from app.services.time import to_company_tz_str
    from app.models import LeadStatus

    wb = Workbook()
    # Bucket leads by campaign name (None → "بدون حملة").
    buckets = {}
    for L in leads:
        key = L.campaign.name if L.campaign else "بدون حملة"
        buckets.setdefault(key, []).append(L)

    # ── Summary sheet ────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "ملخص الحملات"
    _excel_styled_header(
        ws0, "ملخص العملاء المحتملين حسب الحملة",
        company.name, f"عدد الحملات: {len(buckets)}",
    )
    statuses = list(LeadStatus)
    summary_headers = ["الحملة"] + [s.label_ar for s in statuses] + ["الإجمالي"]
    row = 5
    for col, h in enumerate(summary_headers, 1):
        c = ws0.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2540")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws0.column_dimensions["A"].width = 28
    for i in range(len(statuses)):
        ws0.column_dimensions[ws0.cell(row=1, column=2 + i).column_letter].width = 14
    ws0.column_dimensions[ws0.cell(row=1, column=2 + len(statuses)).column_letter].width = 14

    row += 1
    for camp_name in sorted(buckets.keys()):
        camp_leads = buckets[camp_name]
        counts = {s: 0 for s in statuses}
        for L in camp_leads:
            counts[L.status] = counts.get(L.status, 0) + 1
        vals = [camp_name] + [counts[s] for s in statuses] + [len(camp_leads)]
        for col, v in enumerate(vals, 1):
            cell = ws0.cell(row=row, column=col, value=v)
            cell.alignment = Alignment(vertical="center",
                                        horizontal="center")
        # Highlight campaign name cell.
        ws0.cell(row=row, column=1).font = Font(bold=True)
        # Highlight total.
        ws0.cell(row=row, column=len(vals)).font = Font(bold=True)
        row += 1

    # ── One sheet per campaign ──────────────────────────────────────
    headers = [
        "#", "الاسم", "الهاتف", "البريد", "الخدمة",
        "الحالة", "المسؤول", "القيمة المتوقعة",
        "تاريخ الإنشاء", "اجتماع قادم",
    ]
    for camp_name in sorted(buckets.keys()):
        # Sort inside a campaign by status then created_at desc.
        # Excel caps sheet names at 31 chars — truncate defensively.
        title = (camp_name or "بدون حملة")[:31]
        # Also strip characters Excel forbids in sheet names.
        for ch in r'[]:*?/\\':
            title = title.replace(ch, " ")
        ws = wb.create_sheet(title=title)
        camp_leads = sorted(
            buckets[camp_name],
            key=lambda L: (
                list(statuses).index(L.status) if L.status in statuses
                else len(statuses),
                L.created_at or datetime.min,
            ),
        )
        _excel_styled_header(
            ws, f"حملة: {camp_name}",
            company.name, f"عدد العملاء: {len(camp_leads)}",
        )
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0A2540")
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        widths = [4, 22, 18, 26, 22, 14, 18, 14, 18, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=i).column_letter].width = w
        r = 6
        for i, L in enumerate(camp_leads, 1):
            rep = L.assigned_to.full_name if L.assigned_to else ""
            status_lbl = L.status.label_ar if L.status else ""
            expected = float(L.expected_value or 0)
            meeting_str = (to_company_tz_str(
                L.next_meeting, L.company, "%Y-%m-%d %H:%M"
            ) or "" if L.next_meeting else "")
            created_str = (to_company_tz_str(
                L.created_at, L.company, "%Y-%m-%d %H:%M"
            ) or "" if L.created_at else "")
            vals = [
                i, L.client_name or "", L.phone or "", L.email or "",
                L.service_needed or "", status_lbl, rep, expected,
                created_str, meeting_str,
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.alignment = Alignment(vertical="top",
                                            wrap_text=True)
                if col == 8:
                    cell.number_format = "#,##0.00"
            r += 1
        # Totals row for the campaign.
        if camp_leads:
            ws.cell(row=r, column=7,
                     value="إجمالي القيمة المتوقعة:").font = Font(bold=True)
            total = sum(float(L.expected_value or 0) for L in camp_leads)
            c = ws.cell(row=r, column=8, value=total)
            c.number_format = "#,##0.00"
            c.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
