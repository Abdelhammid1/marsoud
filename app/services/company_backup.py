"""MARSOUD — full-company Excel backup.

Generates a single .xlsx workbook with one sheet per major
company-scoped table. Designed for Abdelhamid's workflow: dump
everything, rebuild the chart of accounts, then manually re-key from
the Excel into the new tree.

Public surface:

  build_company_workbook(company_id) -> BytesIO
      Returns an in-memory xlsx ready to send_file.

  workbook_filename(company) -> str
      Suggested filename including company name + timestamp.

Coverage (one sheet per item):
  - Company info
  - Chart of accounts
  - Journal entries + lines (joined view, one row per line)
  - Customers, Vendors, Products, Product variants
  - Invoices + invoice items, Payments
  - Vendor bills
  - Warehouses, Stock balances
  - Employees, Departments, Payroll runs + lines
  - Leave types, Leave balances, Leave requests, Attendance exceptions
  - Projects, Tasks, Leads
  - Fixed assets

Each sheet uses the same header style (navy fill + white bold).
Enum columns export `.value`; FK ids expand to a friendly name where
the relationship is loadable.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app import db
# MARSOUD-CURRENCY-AR — xlsx cells are built in Python, no Jinja filter.
from app.services.currency import currency_name_ar
from app.models import (
    Company, Account, JournalEntry, JournalLine,
    Customer, Vendor, Product,
    Invoice, InvoiceItem, Payment,
    VendorBill,
    Employee, PayrollRun, PayrollLine,
    Project, Task, Lead,
    LeaveType, LeaveBalance, LeaveRequest, AttendanceException,
    FixedAsset,
)
from app.models.inventory import Warehouse, ProductVariant, StockBalance
from app.models.department import Department


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="0A2540")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fmt(value):
    """Render a value for an Excel cell. Enums → .value, datetimes → str,
    None → empty string. Numbers + strings pass through."""
    if value is None:
        return ""
    if hasattr(value, "value") and hasattr(value.__class__, "__members__"):
        return value.value
    if isinstance(value, (datetime,)):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def _write_sheet(ws, headers, rows):
    """Apply header style + write the data rows + autosize columns."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=_fmt(val))
    # Best-effort column widths (capped so insanely long values don't blow it up)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for r in range(2, min(len(rows) + 2, 200)):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                max_len = max(max_len, min(50, len(str(v))))
        ws.column_dimensions[letter].width = max(12, min(50, max_len + 2))
    ws.freeze_panes = "A2"


# ─── Per-sheet builders ─────────────────────────────────────────────────
def _sheet_company(wb, company):
    ws = wb.create_sheet("الشركة")
    _write_sheet(ws,
        ["البيان", "القيمة"],
        [
            ("الاسم", company.name),
            ("العملة الأساسية", currency_name_ar(company.base_currency)),
            ("الرقم الضريبي", getattr(company, "tax_number", "") or ""),
            ("الحالة", getattr(company, "status", "ACTIVE") or "ACTIVE"),
            ("تم الإنشاء", company.created_at),
            ("تاريخ التصدير", datetime.utcnow()),
        ],
    )


def _sheet_accounts(wb, cid):
    rows = Account.query.filter_by(company_id=cid).order_by(Account.code).all()
    ws = wb.create_sheet("شجرة الحسابات")
    _write_sheet(ws,
        ["كود", "الاسم", "الاسم بالعربية", "النوع", "الطرف الطبيعي",
         "كود الأب", "نشط"],
        [(a.code, a.name, a.name_ar or "", a.type, a.normal_side,
          a.parent.code if a.parent else "", a.is_active) for a in rows],
    )


def _sheet_journal_lines(wb, cid):
    """One row per LINE so the workbook is dumpable into a SQL importer
    or a pivot. Includes the parent entry's number/date/description for
    each line — avoids the user having to join sheets manually."""
    lines = db.session.query(JournalLine).join(JournalEntry).filter(
        JournalEntry.company_id == cid
    ).order_by(JournalEntry.date, JournalEntry.id, JournalLine.id).all()
    ws = wb.create_sheet("القيود اليومية")
    _write_sheet(ws,
        ["رقم القيد", "التاريخ", "العملة", "البيان", "مرجع",
         "نوع المصدر", "ID المصدر",
         "كود الحساب", "اسم الحساب", "مدين", "دائن",
         "مدين (أساسي)", "دائن (أساسي)"],
        [(l.entry.number, l.entry.date, l.entry.currency,
          l.entry.description or "", l.entry.reference or "",
          l.entry.source_type or "", l.entry.source_id or "",
          l.account.code if l.account else "",
          l.account.name if l.account else "",
          float(l.debit or 0), float(l.credit or 0),
          float(l.debit_base or 0), float(l.credit_base or 0)) for l in lines],
    )


def _sheet_customers(wb, cid):
    rows = Customer.query.filter_by(company_id=cid).order_by(Customer.name).all()
    ws = wb.create_sheet("العملاء")
    _write_sheet(ws,
        ["الاسم", "الإيميل", "الموبايل", "الرقم الضريبي",
         "نشط", "نسبة العمولة", "تم الإنشاء"],
        [(c.name, c.email or "", c.phone or "", c.tax_number or "",
          c.is_active, float(c.commission_rate or 0), c.created_at)
         for c in rows],
    )


def _sheet_vendors(wb, cid):
    rows = Vendor.query.filter_by(company_id=cid).order_by(Vendor.name).all()
    ws = wb.create_sheet("الموردين")
    _write_sheet(ws,
        ["الاسم", "الإيميل", "الموبايل", "الرقم الضريبي", "نشط", "تم الإنشاء"],
        [(v.name, v.email or "", v.phone or "", v.tax_number or "",
          v.is_active, v.created_at) for v in rows],
    )


def _sheet_products(wb, cid):
    rows = Product.query.filter_by(company_id=cid).order_by(Product.name).all()
    ws = wb.create_sheet("المنتجات")
    _write_sheet(ws,
        ["الاسم", "الوصف", "السعر الافتراضي", "نسبة الضريبة %",
         "SKU", "نشط", "متعقَّب"],
        [(p.name, p.description or "", float(p.default_price or 0),
          float(p.default_tax_rate or 0), p.sku or "",
          p.is_active, getattr(p, "is_tracked", False)) for p in rows],
    )


def _sheet_variants(wb, cid):
    rows = ProductVariant.query.filter_by(company_id=cid).order_by(ProductVariant.sku).all()
    if not rows:
        return
    ws = wb.create_sheet("أصناف المنتجات")
    _write_sheet(ws,
        ["SKU", "الباركود", "اسم الصنف", "اسم المنتج",
         "تكلفة الوحدة", "حد إعادة الطلب", "نشط"],
        [(v.sku or "", v.barcode or "", v.name or "",
          v.product.name if v.product else "",
          float(v.unit_cost or 0), float(v.reorder_level or 0),
          v.is_active) for v in rows],
    )


def _sheet_invoices(wb, cid):
    rows = Invoice.query.filter_by(company_id=cid).order_by(Invoice.issue_date).all()
    ws = wb.create_sheet("فواتير المبيعات")
    _write_sheet(ws,
        ["الرقم", "العميل", "تاريخ الإصدار", "تاريخ الاستحقاق", "الحالة",
         "العملة", "الإجمالي", "المدفوع", "المتبقي", "ملاحظات"],
        [(i.number,
          i.customer.name if i.customer else "",
          i.issue_date, i.due_date, i.status,
          getattr(i, "currency", ""),
          float(i.total or 0), float(i.paid_amount or 0),
          float((i.total or 0) - (i.paid_amount or 0)),
          getattr(i, "notes", "") or "")
         for i in rows],
    )
    # Items go to their own sheet
    items = []
    for inv in rows:
        for it in inv.items:
            items.append((
                inv.number,
                it.product.name if it.product else "",
                it.description or "",
                float(it.quantity or 0),
                float(it.unit_price or 0),
                float(it.line_total or 0),
            ))
    if items:
        ws2 = wb.create_sheet("بنود فواتير المبيعات")
        _write_sheet(ws2,
            ["رقم الفاتورة", "المنتج", "البيان", "الكمية", "سعر الوحدة", "الإجمالي"],
            items,
        )


def _sheet_payments(wb, cid):
    rows = Payment.query.join(Invoice).filter(
        Invoice.company_id == cid
    ).order_by(Payment.payment_date).all()
    if not rows:
        return
    ws = wb.create_sheet("المدفوعات")
    _write_sheet(ws,
        ["تاريخ الدفع", "رقم الفاتورة", "العميل", "المبلغ", "الطريقة"],
        [(p.payment_date,
          p.invoice.number if p.invoice else "",
          p.invoice.customer.name if p.invoice and p.invoice.customer else "",
          float(p.amount or 0),
          (p.payment_method.name if getattr(p, "payment_method", None)
           else (p.method or "")))
         for p in rows],
    )


def _sheet_vendor_bills(wb, cid):
    rows = VendorBill.query.filter_by(company_id=cid).order_by(VendorBill.issue_date).all()
    if not rows:
        return
    ws = wb.create_sheet("فواتير الموردين")
    _write_sheet(ws,
        ["الرقم", "المورد", "تاريخ الإصدار", "تاريخ الاستحقاق", "الحالة",
         "الإجمالي", "المدفوع", "المتبقي"],
        [(b.number,
          b.vendor.name if b.vendor else "",
          b.issue_date, b.due_date, b.status,
          float(b.total or 0), float(b.paid_amount or 0),
          float((b.total or 0) - (b.paid_amount or 0)))
         for b in rows],
    )


def _sheet_warehouses(wb, cid):
    rows = Warehouse.query.filter_by(company_id=cid).all()
    if not rows:
        return
    ws = wb.create_sheet("المخازن")
    _write_sheet(ws,
        ["كود", "الاسم", "افتراضي", "نشط"],
        [(w.code, w.name, w.is_default, w.is_active) for w in rows],
    )


def _sheet_stock(wb, cid):
    rows = StockBalance.query.join(ProductVariant).filter(
        ProductVariant.company_id == cid
    ).all()
    if not rows:
        return
    ws = wb.create_sheet("أرصدة المخزون")
    _write_sheet(ws,
        ["SKU", "اسم الصنف", "المخزن", "الكمية", "تكلفة الوحدة"],
        [(b.variant.sku if b.variant else "",
          b.variant.name if b.variant else "",
          b.warehouse.name if b.warehouse else "",
          float(b.quantity or 0), float(b.unit_cost or 0))
         for b in rows],
    )


def _sheet_employees(wb, cid):
    rows = Employee.query.filter_by(company_id=cid).order_by(Employee.name).all()
    ws = wb.create_sheet("الموظفين")
    _write_sheet(ws,
        ["رقم الموظف", "الاسم", "الإيميل", "المسمى الوظيفي",
         "نوع العقد", "الحالة", "تاريخ التعيين", "الراتب الأساسي"],
        [(e.employee_number or "", e.name, e.email or "",
          getattr(e, "job_title", "") or "",
          getattr(e, "contract_type", ""), e.status,
          getattr(e, "start_date", ""),
          float(getattr(e, "basic_salary", 0) or 0))
         for e in rows],
    )


def _sheet_departments(wb, cid):
    rows = Department.query.filter_by(company_id=cid).all()
    if not rows:
        return
    ws = wb.create_sheet("الأقسام")
    _write_sheet(ws,
        ["الاسم", "الوصف", "نشط"],
        [(d.name, getattr(d, "description", "") or "",
          getattr(d, "is_active", True)) for d in rows],
    )


def _sheet_payroll(wb, cid):
    runs = PayrollRun.query.filter_by(company_id=cid).order_by(
        PayrollRun.period_year, PayrollRun.period_month
    ).all()
    if not runs:
        return
    ws = wb.create_sheet("كشوف الرواتب")
    _write_sheet(ws,
        ["السنة", "الشهر", "الحالة", "تاريخ الإنشاء", "الإجمالي"],
        [(r.period_year, r.period_month,
          getattr(r, "status", "") or "",
          r.created_at,
          float(getattr(r, "total_amount", 0) or 0))
         for r in runs],
    )
    lines = PayrollLine.query.join(PayrollRun).filter(
        PayrollRun.company_id == cid
    ).all()
    if lines:
        ws2 = wb.create_sheet("تفاصيل الرواتب")
        _write_sheet(ws2,
            ["السنة", "الشهر", "الموظف", "الأساسي", "البدلات",
             "الخصومات", "الصافي"],
            [(l.run.period_year if l.run else "",
              l.run.period_month if l.run else "",
              l.employee.name if l.employee else "",
              float(getattr(l, "basic_salary", 0) or 0),
              float(getattr(l, "allowances_total", 0) or 0),
              float(getattr(l, "deductions_total", 0) or 0),
              float(getattr(l, "net_pay", 0) or 0))
             for l in lines],
        )


def _sheet_leaves(wb, cid):
    lt = LeaveType.query.filter_by(company_id=cid).all()
    if lt:
        ws = wb.create_sheet("أنواع الإجازات")
        _write_sheet(ws,
            ["الاسم", "تراكم/شهر", "الحد الأقصى", "مدفوع", "نشط"],
            [(t.name, float(t.accrual_per_month or 0),
              float(t.max_balance or 0), t.is_paid, t.is_active)
             for t in lt],
        )
    balances = LeaveBalance.query.join(Employee).filter(
        Employee.company_id == cid
    ).all()
    if balances:
        ws = wb.create_sheet("أرصدة الإجازات")
        _write_sheet(ws,
            ["الموظف", "نوع الإجازة", "السنة", "الرصيد", "المستخدم", "المتبقي"],
            [(b.employee.name if b.employee else "",
              b.leave_type.name if b.leave_type else "",
              b.year, float(b.balance_days or 0),
              float(b.used_days or 0), float(b.remaining_days or 0))
             for b in balances],
        )
    reqs = LeaveRequest.query.filter_by(company_id=cid).order_by(
        LeaveRequest.start_date.desc()
    ).all()
    if reqs:
        ws = wb.create_sheet("طلبات الإجازة")
        _write_sheet(ws,
            ["الموظف", "نوع الإجازة", "من", "إلى", "عدد الأيام", "الحالة", "السبب"],
            [(r.employee.name if r.employee else "",
              r.leave_type.name if r.leave_type else "",
              r.start_date, r.end_date,
              float(r.days_count or 0), r.status, r.reason or "")
             for r in reqs],
        )
    exc = AttendanceException.query.filter_by(company_id=cid).order_by(
        AttendanceException.date.desc()
    ).all()
    if exc:
        ws = wb.create_sheet("استثناءات الحضور")
        _write_sheet(ws,
            ["الموظف", "التاريخ", "النوع", "ساعات", "ملاحظة"],
            [(e.employee.name if e.employee else "",
              e.date, e.type,
              float(e.duration_hours or 0) if e.duration_hours else "",
              e.note or "") for e in exc],
        )


def _sheet_projects_tasks(wb, cid):
    projects = Project.query.filter_by(company_id=cid).order_by(Project.name).all()
    if projects:
        ws = wb.create_sheet("المشاريع")
        _write_sheet(ws,
            ["الاسم", "العميل", "النوع", "المدير", "تاريخ البداية",
             "تاريخ النهاية", "الحالة"],
            [(p.name,
              p.customer.name if p.customer else "",
              p.type or "",
              p.manager.full_name if p.manager else "",
              p.start_date, p.end_date, p.status)
             for p in projects],
        )
    tasks = Task.query.filter_by(company_id=cid).order_by(Task.created_at).all()
    if tasks:
        ws = wb.create_sheet("المهام")
        _write_sheet(ws,
            ["العنوان", "المشروع", "المسؤول", "الأولوية", "الحالة",
             "Deadline", "تم الإنشاء"],
            [(t.title,
              t.project.name if t.project else "",
              t.assigned_to.full_name if getattr(t, "assigned_to", None) else "",
              t.priority, t.status, t.deadline, t.created_at)
             for t in tasks],
        )


def _sheet_leads(wb, cid):
    rows = Lead.query.filter_by(company_id=cid).order_by(Lead.created_at).all()
    if not rows:
        return
    ws = wb.create_sheet("العملاء المحتملين")
    _write_sheet(ws,
        ["الاسم", "الإيميل", "الموبايل", "الخدمة المطلوبة", "الحالة",
         "المسؤول", "القيمة المتوقعة", "تم الإنشاء"],
        [(l.client_name, l.email or "", l.phone or "",
          getattr(l, "service_needed", "") or "",
          l.status,
          l.assigned_to.full_name if getattr(l, "assigned_to", None) else "",
          float(l.expected_value or 0), l.created_at)
         for l in rows],
    )


def _sheet_assets(wb, cid):
    rows = FixedAsset.query.filter_by(company_id=cid).all()
    if not rows:
        return
    ws = wb.create_sheet("الأصول الثابتة")
    _write_sheet(ws,
        ["الاسم", "التكلفة", "قيمة الخردة", "العمر (سنوات)",
         "الإهلاك المتراكم", "تم التخريد"],
        [(a.name, float(a.cost or 0),
          float(getattr(a, "salvage_value", 0) or 0),
          getattr(a, "useful_life_years", 0) or 0,
          float(getattr(a, "accumulated_depreciation", 0) or 0),
          getattr(a, "is_disposed", False))
         for a in rows],
    )


# ─── Orchestrator ───────────────────────────────────────────────────────
def build_company_workbook(company_id):
    """Build and return an in-memory xlsx (BytesIO) covering every
    company-scoped table. Caller is responsible for streaming via send_file.
    """
    company = db.session.get(Company, company_id)
    if company is None:
        raise ValueError(f"Company {company_id} not found")

    wb = Workbook()
    # Drop the default first sheet — we name our own.
    wb.remove(wb.active)

    _sheet_company(wb, company)
    _sheet_accounts(wb, company_id)
    _sheet_journal_lines(wb, company_id)
    _sheet_customers(wb, company_id)
    _sheet_vendors(wb, company_id)
    _sheet_products(wb, company_id)
    _sheet_variants(wb, company_id)
    _sheet_invoices(wb, company_id)
    _sheet_payments(wb, company_id)
    _sheet_vendor_bills(wb, company_id)
    _sheet_warehouses(wb, company_id)
    _sheet_stock(wb, company_id)
    _sheet_employees(wb, company_id)
    _sheet_departments(wb, company_id)
    _sheet_payroll(wb, company_id)
    _sheet_leaves(wb, company_id)
    _sheet_projects_tasks(wb, company_id)
    _sheet_leads(wb, company_id)
    _sheet_assets(wb, company_id)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def workbook_filename(company):
    """Suggested download filename — slug-ish company name + timestamp."""
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M")
    # Keep only chars Excel + most filesystems handle. Strip path bits.
    safe_name = "".join(
        ch if (ch.isalnum() or ch in "_-") else "_"
        for ch in (company.name or "company")
    )[:50]
    return f"marsoud-backup-{safe_name}-{ts}.xlsx"
